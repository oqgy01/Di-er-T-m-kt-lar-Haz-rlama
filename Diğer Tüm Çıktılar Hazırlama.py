#Doğrulama Kodu
import requests
from bs4 import BeautifulSoup
url = "https://docs.google.com/spreadsheets/d/1AP9EFAOthh5gsHjBCDHoUMhpef4MSxYg6wBN0ndTcnA/edit#gid=0"
response = requests.get(url)
html_content = response.content
soup = BeautifulSoup(html_content, "html.parser")
first_cell = soup.find("td", {"class": "s2"}).text.strip()
if first_cell != "Aktif":
    exit()
first_cell = soup.find("td", {"class": "s1"}).text.strip()
print(first_cell)


import pandas as pd
from openpyxl import load_workbook
from colorama import init, Fore, Style
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
from openpyxl.worksheet.table import Table, TableStyleInfo
import shutil
import zipfile
import os
import datetime
import asyncio
import aiohttp
from concurrent.futures import ThreadPoolExecutor
from tqdm import tqdm
import gc
pd.options.mode.chained_assignment = None
init(autoreset=True)

print(" ")
print(Fore.BLUE + "2500 TL Üzeri Siparişler - Gönderilmesi Gereken Siparişler - Faturasız Siparişler")



print(" ")

bin_tl_uzeri_siparisler = input("2500 TL Üzeri Siparişler Hazırlansın mı? (E/H): ").strip().upper()
gonderilmesi_gereken_siparisler = input("Gönderilmesi Gereken Siparişler Hazırlansın mı? (E/H): ").strip().upper()
faturasiz_siparisler = input("Faturasız Siparişler Hazırlansın mı? (E/H): ").strip().upper()

print(" ")
print("Oturum Açma Başarılı Oldu")
print(" /﹋\ ")
print("(҂`_´)")
print("<,︻╦╤─ ҉ - -")
print("/﹋\\")
print("(Kod Bekçisi)")
print("Mustafa ARI")
print(" ")




#region 2500 TL Üzeri Siparişler

if bin_tl_uzeri_siparisler == "E":



    async def download_file(session, url, index):
        async with session.get(url) as response:
            if response.status == 200:
                file_name = f"link_{index}.xlsx"
                content = await response.read()
                with open(file_name, "wb") as file:
                    file.write(content)
            else:
                print(f"Hata! {url} dosyası indirilemedi.")

    async def main():
        base_url = "https://task.haydigiy.com/FaprikaOrderXls/CXH93A/"
        urls = [f"{base_url}{i}/" for i in range(1, 2)]

        async with aiohttp.ClientSession() as session:
            tasks = [download_file(session, url, index + 1) for index, url in enumerate(urls)]
            await asyncio.gather(*tasks)

        # Excel dosyalarını birleştirip verileri çıkartma
        data_frames = []
        for i in range(1, 2):
            file_name = f"link_{i}.xlsx"
            df = pd.read_excel(file_name)
            data_frames.append(df)

        combined_df = pd.concat(data_frames, ignore_index=True)

        # İstenen sütunları seçme
        selected_columns = ["Id", "OdemeTipi", "SiparisDurumu", "TeslimatTelefon", "Barkod", "Adet", "TeslimatEPostaAdresi", "SiparisToplam", "Varyant", "UrunAdi", "KargoTakipNumarasi", "KargoFirmasi"]
        final_df = combined_df[selected_columns]
        


        # "TeslimatTelefon" sütununda replace yapma
        combined_df['TeslimatTelefon'] = combined_df['TeslimatTelefon'].replace(r'[()/-]', '', regex=True)



        # Virgül ve sonrasını kaldırarak stringi sayıya çevirme işlemi
        def clean_and_convert(value):
            try:
                cleaned_value = value.split(',')[0]  # Virgülden önceki kısmı al
                numeric_value = float(cleaned_value)  # Düzenlenmiş veriyi sayıya çevir
                return numeric_value
            except ValueError:
                return None  # Dönüşüm başarısızsa veya veri boşsa None döndür


        



        # "Adet" ve "SiparisToplam" sütunlarını temizleme ve dönüştürme
        final_df['Adet'] = final_df['Adet'].apply(clean_and_convert)
        final_df['SiparisToplam'] = final_df['SiparisToplam'].apply(clean_and_convert)

        # Düzenlenmiş verileri mevcut dosyanın üzerine kaydetme
        final_df.to_excel("birlesik_excel.xlsx", index=False)
       

        # Birleştirilmiş verileri yeni bir Excel dosyasına kaydetme
        yeni_dosya_adi = "birlesik_excel.xlsx"
        final_df.to_excel(yeni_dosya_adi, index=False)
      


        # İndirilen dosyaları silme
        for i in range(1, 11):
            file_name = f"link_{i}.xlsx"
            if os.path.exists(file_name):
                os.remove(file_name)
                

    if __name__ == "__main__":
        asyncio.run(main())




    






    google_sheet_url = "https://docs.google.com/spreadsheets/d/1FJwRFD6ikSsy3uGFRiKp94Iaj1Jd5xerEzJfxJgS1f8/gviz/tq?tqx=out:csv"

    try:
        google_df = pd.read_csv(google_sheet_url)
        google_excel_file = "Hariç Tutulacak Sipariş Numaraları.xlsx"
        google_df.to_excel(google_excel_file, index=False)
    except requests.exceptions.RequestException as e:
        pass



    def main():
        while True:      
            
            # İki Excel dosyasını okuyun
            birlesik_excel = pd.read_excel("birlesik_excel.xlsx")
            haric_excel = pd.read_excel("Hariç Tutulacak Sipariş Numaraları.xlsx")

            # İlk sütunlara göre birleşik Excel verisinden, haric Excel verisinde bulunan satırları filtreleyin
            birlesik_excel = birlesik_excel[~birlesik_excel['Id'].isin(haric_excel['Id'])]

            # Sonucu mevcut "birlesik_excel.xlsx" dosyasına kaydedin (var olan dosyanın üstüne yazacak)
            birlesik_excel.to_excel("birlesik_excel.xlsx", index=False)
            break
    if __name__ == "__main__":
        main()





    file_path = "Hariç Tutulacak Sipariş Numaraları.xlsx"

    if os.path.exists(file_path):
        os.remove(file_path)










    google_sheet_url = "https://docs.google.com/spreadsheets/d/1zpRV6J8ztpTIKVM0-0wTgfhxZUOPs8Dq/gviz/tq?tqx=out:csv"

    try:
        google_df = pd.read_csv(google_sheet_url)
        google_excel_file = "2500 TL Kurgu.xlsx"
        google_df.to_excel(google_excel_file, index=False)
    except requests.exceptions.RequestException as e:
        print("İstek sırasında bir hata oluştu:", e)










    # İlgili dosyanın adını tanımlayın
    birlesik_excel_adi = "birlesik_excel.xlsx"

    # Veriyi yükleyin
    birlesik_df = pd.read_excel(birlesik_excel_adi)

    # 'TeslimatTelefon' sütunundaki '+' karakterini temizle
    birlesik_df['TeslimatTelefon'] = birlesik_df['TeslimatTelefon'].apply(lambda x: x.replace('+', '') if isinstance(x, str) else x)
    birlesik_df['TeslimatTelefon'] = birlesik_df['TeslimatTelefon'].apply(lambda x: x.replace('-', '') if isinstance(x, str) else x)
    birlesik_df['TeslimatTelefon'] = birlesik_df['TeslimatTelefon'].apply(lambda x: x.replace('(', '') if isinstance(x, str) else x)
    birlesik_df['TeslimatTelefon'] = birlesik_df['TeslimatTelefon'].apply(lambda x: x.replace(')', '') if isinstance(x, str) else x)
    birlesik_df['TeslimatTelefon'] = birlesik_df['TeslimatTelefon'].apply(lambda x: x.replace(' ', '') if isinstance(x, str) else x)
    birlesik_df['TeslimatTelefon'] = birlesik_df['TeslimatTelefon'].apply(lambda x: x.replace('/', '') if isinstance(x, str) else x)

    # Sonuçları aynı dosyanın üzerine kaydedin
    birlesik_df.to_excel(birlesik_excel_adi, index=False)

    










    # İlgili dosyanın adını tanımlayın
    birlesik_excel_adi = "birlesik_excel.xlsx"

    # Veriyi yükleyin
    birlesik_df = pd.read_excel(birlesik_excel_adi)

    # "TeslimatTelefon" sütunundaki tüm değerleri sayıya çevirme
    birlesik_df["TeslimatTelefon"] = birlesik_df["TeslimatTelefon"].apply(lambda x: pd.to_numeric(x, errors='coerce') if x else x)

    # Düzeltilmiş veriyi aynı dosyanın üzerine kaydedin
    birlesik_df.to_excel(birlesik_excel_adi, index=False)

    






















    # İlgili dosyaların adlarını tanımlayın
    birlesik_excel_adi = "birlesik_excel.xlsx"
    kurgu_excel_adi = "2500 TL Kurgu.xlsx"

    # Verileri yükleyin
    birlesik_df = pd.read_excel(birlesik_excel_adi)
    kurgu_df = pd.read_excel(kurgu_excel_adi)

    # Veri eşleştirmesi yapın
    for index, row in birlesik_df.iterrows():
        telefon_numarasi = row['TeslimatTelefon']
        kurgu_satir = kurgu_df[kurgu_df.iloc[:, 0] == telefon_numarasi]
        if not kurgu_satir.empty:
            birlesik_df.at[index, 'Durum'] = kurgu_satir.iloc[0, 1]  # İkinci sütunun değerini alın
        else:
            birlesik_df.at[index, 'Durum'] = "Teyit İçin Ara"  # Eşleşme yoksa "Teyit İçin Ara" yaz

    # Sonuçları aynı dosyanın üzerine kaydedin
    birlesik_df.to_excel(birlesik_excel_adi, index=False)

    



    kurgu_excel_adi = "2500 TL Kurgu.xlsx"

    # Dosyayı sil
    if os.path.exists(kurgu_excel_adi):
        os.remove(kurgu_excel_adi)
    else:
        print(f"{kurgu_excel_adi} dosyası zaten mevcut değil.")








    # İlgili dosyaların adlarını tanımlayın
    birlesik_excel_adi = "birlesik_excel.xlsx"
    kaydedilecek_excel_adi = "2500 TL Üzeri Aranacak Siparişler.xlsx"

    # Verileri yükleyin
    birlesik_df = pd.read_excel(birlesik_excel_adi)

    # "Durum" sütunundaki değeri "Direkt Gönderilir"e eşit olmayan ve
    # "TeslimatEPostaAdresi" sütunundaki değeri "@callcenter" içermeyen satırları ayır
    filtre = (birlesik_df['Durum'] != "Direkt Gönderilir") & (~birlesik_df['TeslimatEPostaAdresi'].str.contains("@callcenter"))
    ayrilanlar_df = birlesik_df[filtre]

    # Ayrılan verileri yeni bir Excel dosyasına kaydet
    ayrilanlar_df.to_excel(kaydedilecek_excel_adi, index=False)

    # "Durum" sütunundaki değeri "Direkt Gönderilir"e eşit olmayan ve
    # "TeslimatEPostaAdresi" sütunundaki değeri "@callcenter" içermeyen satırları sil
    birlesik_df = birlesik_df[~filtre]

    # Kalan verileri aynı dosyanın üzerine kaydet
    birlesik_df.to_excel(birlesik_excel_adi, index=False)




























    # İlgili dosyanın adını tanımlayın
    birlesik_excel_adi = "birlesik_excel.xlsx"

    # Veriyi yükleyin
    birlesik_df = pd.read_excel(birlesik_excel_adi)

    # "Durum" sütununu sil
    birlesik_df.drop(columns=["Durum"], inplace=True)

    # Sonuçları aynı dosyanın üzerine kaydedin
    birlesik_df.to_excel(birlesik_excel_adi, index=False)

    








    # İlgili dosyanın adını tanımlayın
    ayrilan_excel_adi = "2500 TL Üzeri Aranacak Siparişler.xlsx"

    # Veriyi yükleyin
    ayrilan_df = pd.read_excel(ayrilan_excel_adi)

    # "TeslimatTelefon" ve "Durum" sütunlarını kaldır
    ayrilan_df = ayrilan_df.drop(columns=["OdemeTipi", "SiparisDurumu", "Barkod", "Adet", "TeslimatEPostaAdresi", "SiparisToplam", "Varyant", "UrunAdi",])

    # Yalnızca boş sütunlarını koruyun
    ayrilan_df = ayrilan_df[ayrilan_df.columns[ayrilan_df.notna().any()]]

    # Sonuçları aynı dosyanın üzerine kaydedin
    ayrilan_df.to_excel(ayrilan_excel_adi, index=False)

   







    # İlgili dosyanın adını tanımlayın
    ayrilan_excel_adi = "2500 TL Üzeri Aranacak Siparişler.xlsx"

    # Veriyi yükleyin
    ayrilan_df = pd.read_excel(ayrilan_excel_adi)

    # Verileri teke düşür
    ayrilan_df = ayrilan_df.drop_duplicates()

    # Sonuçları aynı dosyanın üzerine kaydedin
    ayrilan_df.to_excel(ayrilan_excel_adi, index=False)

    







    # İlgili dosyanın adını tanımlayın
    excel_adi = "2500 TL Üzeri Aranacak Siparişler.xlsx"

    # Veriyi yükleyin
    df = pd.read_excel(excel_adi)

    # "TeslimatTelefon" sütununu metin olarak biçimlendir
    df['TeslimatTelefon'] = df['TeslimatTelefon'].astype(str)

    # Sonuçları aynı dosyanın üzerine kaydedin
    df.to_excel(excel_adi, index=False)

    













    # İlgili dosyanın adını tanımlayın
    birlesik_excel_adi = "birlesik_excel.xlsx"

    # Veriyi yükleyin
    birlesik_df = pd.read_excel(birlesik_excel_adi)

    # "TeslimatEPostaAdresi" sütununda "@callcenter" içeren satırları sil
    birlesik_df = birlesik_df[~birlesik_df['TeslimatEPostaAdresi'].str.contains("@callcenter")]

    # Sonuçları aynı dosyanın üzerine kaydedin
    birlesik_df.to_excel(birlesik_excel_adi, index=False)

  





    # Kaynak dosya adı
    kaynak_excel = "birlesik_excel.xlsx"

    # Kopya dosya adı (istediğiniz adı ve konumu belirtin)
    kopya_excel = "Kargo Entegrasyonu.xlsx"

    # Dosyayı kopyala
    shutil.copy(kaynak_excel, kopya_excel)





















    # Excel dosyasını oku
    input_file = "birlesik_excel.xlsx"
    df = pd.read_excel(input_file)

    # "OdemeTipi" sütununda "Kapıda Ödeme" olan satırları filtrele
    kapida_odeme_df = df[df["OdemeTipi"] == "Kapıda Ödeme"]

    # "TeslimatTelefon" ile "Id" sütunlarını birleştirerek "BirlesikVeri" sütununu güncelle
    kapida_odeme_df["BirlesikVeri"] = kapida_odeme_df["TeslimatTelefon"].astype(str) + "-" + kapida_odeme_df["Id"].astype(str)

    # "BirlesikVeri" sütunundaki değerlerin tekrar sayılarını hesapla ve yeni bir sütun olarak ekle
    value_counts = kapida_odeme_df["BirlesikVeri"].value_counts()
    kapida_odeme_df["TekrarSayisi"] = kapida_odeme_df["BirlesikVeri"].map(value_counts)

    # "TeslimatTelefon" sütunundaki değerlerin tekrar sayılarını hesapla ve yeni bir sütun olarak ekle
    value_counts2 = kapida_odeme_df["TeslimatTelefon"].value_counts()
    kapida_odeme_df["TekrarSayisi2"] = kapida_odeme_df["TeslimatTelefon"].map(value_counts2)

    # "TekrarSayisi" ve "TekrarSayisi2" sütunlarını karşılaştırarak "TEKRAR" sütunu ekleyin
    kapida_odeme_df["TEKRAR"] = ["TEKRAR" if ts != ts2 else "" for ts, ts2 in zip(kapida_odeme_df["TekrarSayisi"], kapida_odeme_df["TekrarSayisi2"])]

    # Filtreyi kaldır
    df = df[df["OdemeTipi"] != "Kapıda Ödeme"]

    # Güncellenmiş verileri aynı Excel dosyasına kaydet
    df = pd.concat([df, kapida_odeme_df], ignore_index=True)  # Filtrelenmiş verileri ana veri çerçevesine ekleyin
    df.to_excel(input_file, index=False, engine="openpyxl")








    # Excel dosyasını oku
    input_file = "birlesik_excel.xlsx"
    df = pd.read_excel(input_file)

    # "TEKRAR" yazan satırları seç
    tekrar_rows = df[df["TEKRAR"] == "TEKRAR"]

    # "TEKRAR" yazan satırları ayrı bir Excel dosyasına kaydet
    output_file_tekrar = "2500 TL Üzeri Çift Siparişler.xlsx"
    tekrar_rows.to_excel(output_file_tekrar, index=False, engine="openpyxl")

    # "TEKRAR" yazmayan satırları seç ve "TEKRAR" sütununu kaldır
    no_tekrar_rows = df[df["TEKRAR"] != "TEKRAR"]
    no_tekrar_rows = no_tekrar_rows.drop(columns=["TEKRAR"])

    # Güncellenmiş verileri aynı Excel dosyasına kaydet
    no_tekrar_rows.to_excel(input_file, index=False, engine="openpyxl")















    # "2500 TL Üzeri Çift Siparişler.xlsx" dosyasını oku
    tekrar_file = "2500 TL Üzeri Çift Siparişler.xlsx"
    tekrar_df = pd.read_excel(tekrar_file)

    # Sadece "TeslimatTelefon" sütununu tut
    tekrar_df = tekrar_df[["TeslimatTelefon"]]

    # Temizlenmiş "tekrar_satirlar" verilerini aynı dosyaya kaydet
    tekrar_df.to_excel(tekrar_file, index=False, engine="openpyxl")

    # Tekrar satırları temizle
    tekrar_df.drop_duplicates(subset="TeslimatTelefon", inplace=True)

    # "TeslimatTelefon" sütununu sayı biçimine çevir (formatlama)
    try:
        tekrar_df["TeslimatTelefon"] = tekrar_df["TeslimatTelefon"].apply(lambda x: '{:.0f}'.format(x))
    except ValueError:
        pass

    # Temizlenmiş tekrar satırları ayrı bir Excel dosyasına kaydet
    output_cleaned_file = "2500 TL Üzeri Çift Siparişler.xlsx"
    tekrar_df.to_excel(output_cleaned_file, index=False, engine="openpyxl")












    # İlgili dosyanın adını tanımlayın
    ayrilan_excel_adi = "birlesik_excel.xlsx"

    # Veriyi yükleyin
    ayrilan_df = pd.read_excel(ayrilan_excel_adi)

    # "TeslimatTelefon" ve "Durum" sütunlarını kaldır
    ayrilan_df = ayrilan_df.drop(columns=["TekrarSayisi", "TekrarSayisi2", "BirlesikVeri"])

    # Yalnızca boş sütunlarını koruyun
    ayrilan_df = ayrilan_df[ayrilan_df.columns[ayrilan_df.notna().any()]]

    # Sonuçları aynı dosyanın üzerine kaydedin
    ayrilan_df.to_excel(ayrilan_excel_adi, index=False)

 











    # Excel dosyasını okuma
    df = pd.read_excel("birlesik_excel.xlsx")


    # İşlemi gerçekleştiren fonksiyon
    def duplicate_rows(row):
        count = int(row["Adet"])
        return pd.concat([row] * count, axis=1).T

    # Tüm satırları işleme tabi tutma
    new_rows = df.apply(duplicate_rows, axis=1)

    # Yeni veri çerçevesini oluşturma
    new_df = pd.concat(new_rows.tolist(), ignore_index=True)

    # Sadece belirtilen sütunları seçme
    selected_columns = ["Id", "Barkod", "UrunAdi", "Varyant"]
    new_df = new_df[selected_columns]

    # Veriyi yeni bir Excel dosyasına yazma
    new_df.to_excel("birlesik_excel.xlsx", index=False)

   




    # Kaynak dosya adı
    kaynak_excel = "birlesik_excel.xlsx"

    # Kopya dosya adı (istediğiniz adı ve konumu belirtin)
    kopya_excel = "Hazırlanan Sipariş Numaraları.xlsx"

    # Dosyayı kopyala
    shutil.copy(kaynak_excel, kopya_excel)






    url = "https://haydigiy.online/Products/rafkodlari.php"
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")
    table = soup.find("table")
    data = []
    for row in table.find_all("tr"):
        row_data = []
        for cell in row.find_all(["th", "td"]):
            row_data.append(cell.get_text(strip=True))
        data.append(row_data)
    df = pd.DataFrame(data[1:], columns=data[0])
    df.to_excel("Raf Kodu.xlsx", index=False)
      






    # "birlesik_excel.xlsx" ve "Raf Kodu.xlsx" dosyalarını okuma
    sonuc_df = pd.read_excel("birlesik_excel.xlsx")
    google_sheet_df = pd.read_excel("Raf Kodu.xlsx")

    # "birlesik_excel.xlsx" dosyasına yeni bir sütun ekleyerek başlangıçta "Raf Kodu Yok" değerleri ile doldurma
    sonuc_df["GoogleSheetVerisi"] = "Raf Kodu Yok"

    # Her bir "Barkod" değeri için işlem yapma
    for index, row in sonuc_df.iterrows():
        barkod = row["Barkod"]
        
        # "Raf Kodu.xlsx" dosyasında ilgili "Barkod"u arama
        matching_row = google_sheet_df[google_sheet_df.iloc[:, 0] == barkod]
        
        # Eşleşen "Barkod" varsa ve karşılık gelen hücre boş değilse, değeri "GoogleSheetVerisi" sütununa yazma
        if not matching_row.empty and not pd.isnull(matching_row.iloc[0, 2]):
            sonuc_df.at[index, "GoogleSheetVerisi"] = matching_row.iloc[0, 2]

    # "birlesik_excel.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("birlesik_excel.xlsx", index=False)
    






    # "birlesik_excel.xlsx" dosyasını güncelleme
    sonuc_df["GoogleSheetVerisi Kopya"] = sonuc_df["GoogleSheetVerisi"]  # "GoogleSheetVerisi" sütununu kopyala
    sonuc_df["GoogleSheetVerisi Kopya"] = sonuc_df["GoogleSheetVerisi Kopya"].str.split("-", n=1).str[0]  # "-" den sonrasını temizle
    sonuc_df["GoogleSheetVerisi Kopya"] = pd.to_numeric(sonuc_df["GoogleSheetVerisi Kopya"], errors="coerce")  # Sayıya dönüştür
    sonuc_df = sonuc_df.sort_values(by="GoogleSheetVerisi Kopya")  # "GoogleSheetVerisi Kopya" sütununa göre sırala


    sonuc_df.to_excel("birlesik_excel.xlsx", index=False)
    













    # "birlesik_excel.xlsx" ve "Raf Kodu.xlsx" dosyalarını okuma
    sonuc_df = pd.read_excel("birlesik_excel.xlsx")
    google_sheet_df = pd.read_excel("Raf Kodu.xlsx")

    # "birlesik_excel.xlsx" dosyasına yeni bir sütun ekleyerek başlangıçta "Raf Kodu Yok" değerleri ile doldurma
    sonuc_df["Kategori"] = "Raf Kodu Yok"

    # Her bir "Barkod" değeri için işlem yapma
    for index, row in sonuc_df.iterrows():
        barkod = row["Barkod"]
        
        # "Raf Kodu.xlsx" dosyasında ilgili "Barkod"u arama
        matching_row = google_sheet_df[google_sheet_df.iloc[:, 0] == barkod]
        
        # Eşleşen "Barkod" varsa ve karşılık gelen hücre boş değilse, değeri "GoogleSheetVerisi" sütununa yazma
        if not matching_row.empty and not pd.isnull(matching_row.iloc[0, 3]):
            sonuc_df.at[index, "Kategori"] = matching_row.iloc[0, 3]

    # "birlesik_excel.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("birlesik_excel.xlsx", index=False)
  





    # Excel dosyasını oku
    excel_file_path = "birlesik_excel.xlsx"  # Excel dosyasının adını ve yolunu belirtin
    df = pd.read_excel(excel_file_path)

    # Yeni bir sütun ekleyerek işlem sonuçlarını tut
    df["Yeni Kategori"] = ""

    # İç Giyim içeren satırları işle
    innerwear_rows = df[df["Kategori"].str.contains("İç Giyim")]

    # İç Giyim içeren satırları işle
    for index, row in innerwear_rows.iterrows():
        df.loc[index, "Yeni Kategori"] = "İç Giyim"

    # Sonucu kaydet
    output_file_path = "birlesik_excel.xlsx"  # Sonucun kaydedileceği dosya adı ve yolunu belirtin
    df.to_excel(output_file_path, index=False)

    

    # Excel dosyasını oku
    excel_file_path = "birlesik_excel.xlsx"  # Excel dosyasının adını ve yolunu belirtin
    df = pd.read_excel(excel_file_path)

    # "Id" değerlerine göre grupla ve işlemi yap
    grouped = df.groupby("Id")

    # "Id" değerlerine göre "Yeni Kategori" değerini güncelle
    for group_name, group_data in grouped:
        if any(row["Yeni Kategori"] == "İç Giyim" for _, row in group_data.iterrows()):
            df.loc[df["Id"] == group_name, "Yeni Kategori"] = "İç Giyim"

    # Sonucu kaydet
    output_file_path = "birlesik_excel.xlsx"  # Sonucun kaydedileceği dosya adı ve yolunu belirtin
    df.to_excel(output_file_path, index=False)

    







    # Excel dosyasını oku
    excel_file_path = "birlesik_excel.xlsx"  # Ana Excel dosyasının adını ve yolunu belirtin
    df = pd.read_excel(excel_file_path)

    # "Yeni Kategori" değeri "İç Giyim" olan satırları seç
    innerwear_rows = df[df["Yeni Kategori"] == "İç Giyim"]

    # Ayrı Excel dosyasına kaydet
    output_file_path = "İç Giyim.xlsx"  # Ayrı dosyanın adını ve yolunu belirtin
    innerwear_rows.to_excel(output_file_path, index=False)

    # Ana DataFrame'den "İç Giyim" satırları sil
    df = df[df["Yeni Kategori"] != "İç Giyim"]
    df.drop(columns=["Yeni Kategori"], inplace=True)  # "Yeni Kategori" sütununu sil

    # Ana Excel dosyasını güncelle
    df.to_excel(excel_file_path, index=False)

    












    excel_to_delete = "Raf Kodu.xlsx"

    # Excel dosyasını sil
    if os.path.exists(excel_to_delete):
        os.remove(excel_to_delete)
        
    else:
        print("Dosya bulunamadı:", excel_to_delete)



















    old_file_path = "birlesik_excel.xlsx"
    new_file_path = "14.xlsx"

    # Dosyanın adını değiştir
    os.rename(old_file_path, new_file_path)

    







    # Excel dosyasını oku
    excel_file_path = "14.xlsx"  # 14 Excel dosyasının adını ve yolunu belirtin
    df = pd.read_excel(excel_file_path)

    # "GoogleSheetVerisi Kopya" ve "Kategori" sütunlarını sil
    columns_to_drop = ["GoogleSheetVerisi Kopya", "Kategori"]
    df.drop(columns=columns_to_drop, inplace=True)

    # Dosyayı güncelle
    df.to_excel(excel_file_path, index=False)

   





    # Excel dosyasını oku
    excel_file_path = "İç Giyim.xlsx"  # 14 Excel dosyasının adını ve yolunu belirtin
    df = pd.read_excel(excel_file_path)

    # "GoogleSheetVerisi Kopya" ve "Kategori" sütunlarını sil
    columns_to_drop = ["GoogleSheetVerisi Kopya", "Kategori", "Yeni Kategori"]
    df.drop(columns=columns_to_drop, inplace=True)

    # Dosyayı güncelle
    df.to_excel(excel_file_path, index=False)

   



























































    sonuc_df = pd.read_excel("14.xlsx")

    # "14.xlsx" dosyasını güncelleme
    sonuc_df["UrunAdi Kopya"] = sonuc_df["UrunAdi"]  # "UrunAdi" sütununu kopyala
    sonuc_df["UrunAdi Kopya"] = sonuc_df["UrunAdi Kopya"].str.split("-", n=1).str[1]  # "-" den öncesini temizle

    # "UrunAdi Kopya" sütununu "14.xlsx" dosyasına ekleyerek güncelleme
    with pd.ExcelWriter("14.xlsx") as writer:
        sonuc_df.to_excel(writer, index=False)

    








    # "UrunAdi" sütununu en sağına yapıştırma
    sonuc_df["UrunAdiKopya2"] = sonuc_df["UrunAdi"]

    # "UrunAdiKopya2" sütununda " - " dan sonrasını ve son boşluktan öncesini silme
    sonuc_df["UrunAdiKopya2"] = sonuc_df["UrunAdiKopya2"].apply(lambda x: x.split(" - ")[0].rsplit(" ", 1)[-1].strip() if " - " in x else x)

    # "UrunAdiKopya2" sütununu en sağa taşıma
    column_order = list(sonuc_df.columns)
    column_order.remove("UrunAdiKopya2")
    column_order.append("UrunAdiKopya2")
    sonuc_df = sonuc_df[column_order]

    # "14.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("14.xlsx", index=False)
    






    # "UrunAdi" sütununu tablonun sonuna bir kez daha kopyalama ve düzenleme
    sonuc_df["UrunAdiKopya3"] = sonuc_df["UrunAdi"].apply(lambda x: x.split(" - ", 1)[0].strip() if " - " in x else x)

    # "UrunAdiKopya3" sütununu en sağa taşıma
    column_order = list(sonuc_df.columns)
    column_order.remove("UrunAdiKopya3")
    column_order.append("UrunAdiKopya3")
    sonuc_df = sonuc_df[column_order]

    # "14.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("14.xlsx", index=False)
    









    # Verileri birleştirip yeni sütun oluşturma
    sonuc_df["BirlesikVeri"] = sonuc_df["UrunAdi Kopya"] + " - " + sonuc_df["UrunAdiKopya2"] + " - " + sonuc_df["Varyant"]

    # "BirlesikVeri" sütununu en sağa taşıma
    column_order = list(sonuc_df.columns)
    column_order.remove("BirlesikVeri")
    column_order.append("BirlesikVeri")
    sonuc_df = sonuc_df[column_order]

    # "14.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("14.xlsx", index=False)
    

    # "BirlesikVeri" sütunundaki "Beden:" ibaresini çıkarma
    sonuc_df["BirlesikVeri"] = sonuc_df["BirlesikVeri"].str.replace("Beden:", "")

    # "14.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("14.xlsx", index=False)
    





    # Belirtilen sütunları silme
    columns_to_drop = ["UrunAdi", "Varyant", "UrunAdi Kopya", "UrunAdiKopya2"]
    sonuc_df.drop(columns_to_drop, axis=1, inplace=True)

    # "14.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("14.xlsx", index=False)
    






    # "Id" sütununu teke düşürme
    unique_ids = sonuc_df["Id"].drop_duplicates()

    # Yeni bir Excel sayfası oluşturma ve "Id" değerlerini yazma
    with pd.ExcelWriter("14.xlsx", engine="openpyxl", mode="a") as writer:
        unique_ids.to_excel(writer, sheet_name="Unique Ids", index=False)

    












    # Sonuç dosyasını yükle
    file_path = "14.xlsx"
    wb = load_workbook(file_path)
    sheet = wb["Unique Ids"]

    # Başlangıç sütunu ve satırı
    start_column = 2
    start_row = 2

    # Toplam tekrar sayısı ve her tekrardaki numara adedi
    repeat_count = 100
    numbers_per_repeat = 14

    # Verileri ekleme
    for _ in range(repeat_count):
        for num in range(1, numbers_per_repeat + 1):
            sheet.cell(row=start_row, column=start_column, value=num)
            start_row += 1

    # Değişiklikleri kaydetme
    wb.save("14.xlsx")






    # Sonuç dosyasını yükle
    file_path = "14.xlsx"
    wb = load_workbook(file_path)
    sheet = wb["Unique Ids"]

    # Başlangıç sütunu ve satırı
    start_column = 3
    start_row = 2

    # Toplam tekrar sayısı ve her tekrardaki numara adedi
    repeat_count = 14
    numbers_per_repeat = 100

    # Verileri ekleme
    for num in range(1, numbers_per_repeat + 1):
        for repeat in range(repeat_count):
            sheet.cell(row=start_row, column=start_column, value=num)
            start_row += 1

    # Değişiklikleri kaydetme
    wb.save("14.xlsx")

    









    # Sonuç dosyasını yükle
    file_path = "14.xlsx"
    wb = load_workbook(file_path)
    unique_ids_sheet = wb["Unique Ids"]
    main_sheet = wb["Sheet1"]

    # "Id" sütununun verilerini al
    id_column = main_sheet["A"][1:]
    unique_ids_column = unique_ids_sheet["A"][1:]

    # Karşılık gelen değerleri bulup "Sheet1" sayfasının en sağında yeni bir sütuna ekle
    new_column = main_sheet.max_column + 1
    main_sheet.cell(row=1, column=new_column, value="Matching Value")

    for id_cell in id_column:
        id_value = id_cell.value
        for unique_id_cell in unique_ids_column:
            if unique_id_cell.value == id_value:
                matching_value = unique_id_cell.offset(column=1).value
                main_sheet.cell(row=id_cell.row, column=new_column, value=matching_value)
                break

    # Değişiklikleri kaydetme
    wb.save("14.xlsx")

    










    # Sonuç dosyasını yükle
    file_path = "14.xlsx"
    wb = load_workbook(file_path)
    unique_ids_sheet = wb["Unique Ids"]
    main_sheet = wb["Sheet1"]

    # "Id" sütununun verilerini al
    id_column = main_sheet["A"][1:]
    unique_ids_column = unique_ids_sheet["A"][1:]

    # Karşılık gelen değerleri bulup "Sheet1" sayfasının en sağında yeni bir sütuna ekle
    new_column = main_sheet.max_column + 1
    main_sheet.cell(row=1, column=new_column, value="Matching Value (3rd Column)")

    for id_cell in id_column:
        id_value = id_cell.value
        for unique_id_cell in unique_ids_column:
            if unique_id_cell.value == id_value:
                matching_value = unique_id_cell.offset(column=2).value
                main_sheet.cell(row=id_cell.row, column=new_column, value=matching_value)
                break

    # Değişiklikleri kaydetme
    wb.save("14.xlsx")

   









    # Sonuç dosyasını yükle
    file_path = "14.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Sütun başlıklarını değiştir
    new_column_titles = {
        "Id": "SiparişNO",
        "BirlesikVeri": "ÜRÜN",
        "GoogleSheetVerisi": "RAF KODU",
        "UrunAdiKopya3": "ÜRÜN ADI",
        "Matching Value": "KUTU",
        "Matching Value (3rd Column)": "ÇN"
    }

    for col_idx, col_name in enumerate(main_sheet.iter_cols(min_col=1, max_col=main_sheet.max_column), start=1):
        old_title = col_name[0].value
        new_title = new_column_titles.get(old_title, old_title)
        col_name[0].value = new_title

    # Değişiklikleri kaydetme
    wb.save("14.xlsx")

   








    

    # Sonuç dosyasını yükle
    file_path = "14.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Sütunların yeni sıralaması
    new_column_order = [
        "RAF KODU",
        "ÜRÜN",
        "Barkod",
        "KUTU",
        "ÜRÜN ADI",
        "ÇN",
        "SiparişNO"
    ]

    # Yeni bir DataFrame oluştur
    data = main_sheet.iter_rows(min_row=2, values_only=True)
    df = pd.DataFrame(data, columns=[cell.value for cell in main_sheet[1]])

    # Sütunları yeni sıralamaya göre düzenle
    df = df[new_column_order]

    # Mevcut başlıkları güncelle
    for idx, column_name in enumerate(new_column_order, start=1):
        main_sheet.cell(row=1, column=idx, value=column_name)

    # DataFrame verilerini sayfaya yaz
    for r_idx, row in enumerate(df.values, 2):
        for c_idx, value in enumerate(row, 1):
            main_sheet.cell(row=r_idx, column=c_idx, value=value)

    # Değişiklikleri kaydet
    wb.save("14.xlsx")

    







    

    # Sonuç dosyasını yükle
    file_path = "14.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Hücreleri ortala ve ortaya hizala
    for row in main_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")


    # Değişiklikleri kaydet
    wb.save("14.xlsx")

    








    # Sonuç dosyasını yükle
    file_path = "14.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Kenarlık stili oluştur
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Hücreleri kalın yap, yazı fontunu 14 yap ve kenarlık ekle
    for row in main_sheet.iter_rows():
        for cell in row:
            cell.font = Font(bold=True, size=14)
            cell.border = border_style

    # Değişiklikleri kaydet
    wb.save("14.xlsx")

    


    

    # Sonuç dosyasını yükle
    file_path = "14.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # "RAF KODU" sütununu 45 piksel yap
    main_sheet.column_dimensions["C"].width = 45

    # Tüm hücreleri en uygun sütun genişliği olarak ayarla
    for column in main_sheet.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        main_sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Değişiklikleri kaydet
    wb.save("14.xlsx")

    



    # Sonuç dosyasını yükle
    file_path = "14.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # İlk sütunu (A sütunu) 45 piksel genişliğinde yap
    main_sheet.column_dimensions["A"].width = 45
    main_sheet.column_dimensions["C"].width = 14
    main_sheet.column_dimensions["G"].width = 14
    main_sheet.column_dimensions["D"].width = 9
    main_sheet.column_dimensions["F"].width = 5

    # Değişiklikleri kaydet
    wb.save("14.xlsx")

    








    

    # Sonuç dosyasını yükle
    file_path = "14.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Tüm hücrelere "Metni Kaydır" formatını uygula
    for row in main_sheet.iter_rows():
        for cell in row:
            new_alignment = copy(cell.alignment)
            new_alignment.wrap_text = True
            cell.alignment = new_alignment

    # Değişiklikleri kaydet
    wb.save("14.xlsx")

    







    

    # Sonuç dosyasını yükle
    file_path = "14.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Tabloyu oluşturma
    table = Table(displayName="MyTable", ref=main_sheet.dimensions)

    # Tablo stili oluşturma (gri-beyaz)
    style = TableStyleInfo(
        name="TableStyleLight1", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True
    )

    # Tabloya stil atama
    table.tableStyleInfo = style

    # Tabloyu sayfaya ekleme
    main_sheet.add_table(table)

    # Değişiklikleri kaydetme
    wb.save("14.xlsx")

   


    

    def create_bat_files(data, output_folder, batch_size=14):
        batch_count = 1
        batch_data = []
        remaining_data = data

        while len(remaining_data) > 0:
            current_batch = remaining_data[:batch_size]
            batch_data.extend(current_batch)

            bat_file_path = os.path.join(output_folder, f"BAT{batch_count}.bat")
            with open(bat_file_path, "w") as file:
                link = f'start "" https://task.haydigiy.com/admin/order/printorder/?orderId={current_batch[0]}&isPdf=False\n'
                file.write(link)
                file.write('timeout -t 1\n')  # Add the timeout line

                for value in current_batch[1:]:
                    link = f"https://task.haydigiy.com/admin/order/printorder/?orderId={value}&isPdf=False"
                    file.write(f'start "" {link}\n')

            batch_data = []
            remaining_data = remaining_data[batch_size:]
            batch_count += 1

    # Klasör oluştur
    output_folder = "14"
    os.makedirs(output_folder, exist_ok=True)

    # Sonuç dosyasını yükle
    file_path = "14.xlsx"
    wb = load_workbook(file_path)
    unique_ids_sheet = wb["Unique Ids"]

    # "Id" sütunundaki verileri al
    id_column = unique_ids_sheet["A"][1:]

    # Verileri bir listeye dönüştür
    id_values = [cell.value for cell in id_column if cell.value is not None]

    # .bat dosyalarını oluştur ve klasöre kaydet
    create_bat_files(id_values, output_folder)

    # Excel dosyasını klasöre taşı
    shutil.copy(file_path, os.path.join(output_folder, "14.xlsx"))

    gc.collect()

    # Klasör dışında kalan Excel dosyasını sil
    os.remove(file_path)

    




















































    sonuc_df = pd.read_excel("İç Giyim.xlsx")

    # "İç Giyim.xlsx" dosyasını güncelleme
    sonuc_df["UrunAdi Kopya"] = sonuc_df["UrunAdi"]  # "UrunAdi" sütununu kopyala
    sonuc_df["UrunAdi Kopya"] = sonuc_df["UrunAdi Kopya"].str.split("-", n=1).str[1]  # "-" den öncesini temizle

    # "UrunAdi Kopya" sütununu "İç Giyim.xlsx" dosyasına ekleyerek güncelleme
    with pd.ExcelWriter("İç Giyim.xlsx") as writer:
        sonuc_df.to_excel(writer, index=False)

    








    # "UrunAdi" sütununu en sağına yapıştırma
    sonuc_df["UrunAdiKopya2"] = sonuc_df["UrunAdi"]

    # "UrunAdiKopya2" sütununda " - " dan sonrasını ve son boşluktan öncesini silme
    sonuc_df["UrunAdiKopya2"] = sonuc_df["UrunAdiKopya2"].apply(lambda x: x.split(" - ")[0].rsplit(" ", 1)[-1].strip() if " - " in x else x)

    # "UrunAdiKopya2" sütununu en sağa taşıma
    column_order = list(sonuc_df.columns)
    column_order.remove("UrunAdiKopya2")
    column_order.append("UrunAdiKopya2")
    sonuc_df = sonuc_df[column_order]

    # "İç Giyim.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("İç Giyim.xlsx", index=False)
    






    # "UrunAdi" sütununu tablonun sonuna bir kez daha kopyalama ve düzenleme
    sonuc_df["UrunAdiKopya3"] = sonuc_df["UrunAdi"].apply(lambda x: x.split(" - ", 1)[0].strip() if " - " in x else x)

    # "UrunAdiKopya3" sütununu en sağa taşıma
    column_order = list(sonuc_df.columns)
    column_order.remove("UrunAdiKopya3")
    column_order.append("UrunAdiKopya3")
    sonuc_df = sonuc_df[column_order]

    # "İç Giyim.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("İç Giyim.xlsx", index=False)
    









    # Verileri birleştirip yeni sütun oluşturma
    sonuc_df["BirlesikVeri"] = sonuc_df["UrunAdi Kopya"] + " - " + sonuc_df["UrunAdiKopya2"] + " - " + sonuc_df["Varyant"]

    # "BirlesikVeri" sütununu en sağa taşıma
    column_order = list(sonuc_df.columns)
    column_order.remove("BirlesikVeri")
    column_order.append("BirlesikVeri")
    sonuc_df = sonuc_df[column_order]

    # "İç Giyim.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("İç Giyim.xlsx", index=False)
    

    # "BirlesikVeri" sütunundaki "Beden:" ibaresini çıkarma
    sonuc_df["BirlesikVeri"] = sonuc_df["BirlesikVeri"].str.replace("Beden:", "")

    # "İç Giyim.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("İç Giyim.xlsx", index=False)
    





    # Belirtilen sütunları silme
    columns_to_drop = ["UrunAdi", "Varyant", "UrunAdi Kopya", "UrunAdiKopya2"]
    sonuc_df.drop(columns_to_drop, axis=1, inplace=True)

    # "İç Giyim.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("İç Giyim.xlsx", index=False)
    






    # "Id" sütununu teke düşürme
    unique_ids = sonuc_df["Id"].drop_duplicates()

    # Yeni bir Excel sayfası oluşturma ve "Id" değerlerini yazma
    with pd.ExcelWriter("İç Giyim.xlsx", engine="openpyxl", mode="a") as writer:
        unique_ids.to_excel(writer, sheet_name="Unique Ids", index=False)

   











    

    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    sheet = wb["Unique Ids"]

    # Başlangıç sütunu ve satırı
    start_column = 2
    start_row = 2

    # Toplam tekrar sayısı ve her tekrardaki numara adedi
    repeat_count = 100
    numbers_per_repeat = 14

    # Verileri ekleme
    for _ in range(repeat_count):
        for num in range(1, numbers_per_repeat + 1):
            sheet.cell(row=start_row, column=start_column, value=num)
            start_row += 1

    # Değişiklikleri kaydetme
    wb.save("İç Giyim.xlsx")






    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    sheet = wb["Unique Ids"]

    # Başlangıç sütunu ve satırı
    start_column = 3
    start_row = 2

    # Toplam tekrar sayısı ve her tekrardaki numara adedi
    repeat_count = 14
    numbers_per_repeat = 100

    # Verileri ekleme
    for num in range(1, numbers_per_repeat + 1):
        for repeat in range(repeat_count):
            sheet.cell(row=start_row, column=start_column, value=num)
            start_row += 1

    # Değişiklikleri kaydetme
    wb.save("İç Giyim.xlsx")

   









    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    unique_ids_sheet = wb["Unique Ids"]
    main_sheet = wb["Sheet1"]

    # "Id" sütununun verilerini al
    id_column = main_sheet["A"][1:]
    unique_ids_column = unique_ids_sheet["A"][1:]

    # Karşılık gelen değerleri bulup "Sheet1" sayfasının en sağında yeni bir sütuna ekle
    new_column = main_sheet.max_column + 1
    main_sheet.cell(row=1, column=new_column, value="Matching Value")

    for id_cell in id_column:
        id_value = id_cell.value
        for unique_id_cell in unique_ids_column:
            if unique_id_cell.value == id_value:
                matching_value = unique_id_cell.offset(column=1).value
                main_sheet.cell(row=id_cell.row, column=new_column, value=matching_value)
                break

    # Değişiklikleri kaydetme
    wb.save("İç Giyim.xlsx")

    










    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    unique_ids_sheet = wb["Unique Ids"]
    main_sheet = wb["Sheet1"]

    # "Id" sütununun verilerini al
    id_column = main_sheet["A"][1:]
    unique_ids_column = unique_ids_sheet["A"][1:]

    # Karşılık gelen değerleri bulup "Sheet1" sayfasının en sağında yeni bir sütuna ekle
    new_column = main_sheet.max_column + 1
    main_sheet.cell(row=1, column=new_column, value="Matching Value (3rd Column)")

    for id_cell in id_column:
        id_value = id_cell.value
        for unique_id_cell in unique_ids_column:
            if unique_id_cell.value == id_value:
                matching_value = unique_id_cell.offset(column=2).value
                main_sheet.cell(row=id_cell.row, column=new_column, value=matching_value)
                break

    # Değişiklikleri kaydetme
    wb.save("İç Giyim.xlsx")

    









    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Sütun başlıklarını değiştir
    new_column_titles = {
        "Id": "SiparişNO",
        "BirlesikVeri": "ÜRÜN",
        "GoogleSheetVerisi": "RAF KODU",
        "UrunAdiKopya3": "ÜRÜN ADI",
        "Matching Value": "KUTU",
        "Matching Value (3rd Column)": "ÇN"
    }

    for col_idx, col_name in enumerate(main_sheet.iter_cols(min_col=1, max_col=main_sheet.max_column), start=1):
        old_title = col_name[0].value
        new_title = new_column_titles.get(old_title, old_title)
        col_name[0].value = new_title

    # Değişiklikleri kaydetme
    wb.save("İç Giyim.xlsx")

    








    

    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Sütunların yeni sıralaması
    new_column_order = [
        "RAF KODU",
        "ÜRÜN",
        "Barkod",
        "KUTU",
        "ÜRÜN ADI",
        "ÇN",
        "SiparişNO"
    ]

    # Yeni bir DataFrame oluştur
    data = main_sheet.iter_rows(min_row=2, values_only=True)
    df = pd.DataFrame(data, columns=[cell.value for cell in main_sheet[1]])

    # Sütunları yeni sıralamaya göre düzenle
    df = df[new_column_order]

    # Mevcut başlıkları güncelle
    for idx, column_name in enumerate(new_column_order, start=1):
        main_sheet.cell(row=1, column=idx, value=column_name)

    # DataFrame verilerini sayfaya yaz
    for r_idx, row in enumerate(df.values, 2):
        for c_idx, value in enumerate(row, 1):
            main_sheet.cell(row=r_idx, column=c_idx, value=value)

    # Değişiklikleri kaydet
    wb.save("İç Giyim.xlsx")

    







    

    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Hücreleri ortala ve ortaya hizala
    for row in main_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")


    # Değişiklikleri kaydet
    wb.save("İç Giyim.xlsx")

    








    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Kenarlık stili oluştur
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Hücreleri kalın yap, yazı fontunu 14 yap ve kenarlık ekle
    for row in main_sheet.iter_rows():
        for cell in row:
            cell.font = Font(bold=True, size=14)
            cell.border = border_style

    # Değişiklikleri kaydet
    wb.save("İç Giyim.xlsx")

   


    

    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # "RAF KODU" sütununu 45 piksel yap
    main_sheet.column_dimensions["C"].width = 45

    # Tüm hücreleri en uygun sütun genişliği olarak ayarla
    for column in main_sheet.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        main_sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Değişiklikleri kaydet
    wb.save("İç Giyim.xlsx")

    



    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # İlk sütunu (A sütunu) 45 piksel genişliğinde yap
    main_sheet.column_dimensions["A"].width = 45
    main_sheet.column_dimensions["C"].width = 14
    main_sheet.column_dimensions["G"].width = 14
    main_sheet.column_dimensions["D"].width = 9
    main_sheet.column_dimensions["F"].width = 5

    # Değişiklikleri kaydet
    wb.save("İç Giyim.xlsx")

    








    

    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Tüm hücrelere "Metni Kaydır" formatını uygula
    for row in main_sheet.iter_rows():
        for cell in row:
            new_alignment = copy(cell.alignment)
            new_alignment.wrap_text = True
            cell.alignment = new_alignment

    # Değişiklikleri kaydet
    wb.save("İç Giyim.xlsx")

    







    

    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Tabloyu oluşturma
    table = Table(displayName="MyTable", ref=main_sheet.dimensions)

    # Tablo stili oluşturma (gri-beyaz)
    style = TableStyleInfo(
        name="TableStyleLight1", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True
    )

    # Tabloya stil atama
    table.tableStyleInfo = style

    # Tabloyu sayfaya ekleme
    main_sheet.add_table(table)

    # Değişiklikleri kaydetme
    wb.save("İç Giyim.xlsx")

  


    

    def create_bat_files(data, output_folder, batch_size=14):
        batch_count = 1
        batch_data = []
        remaining_data = data

        while len(remaining_data) > 0:
            current_batch = remaining_data[:batch_size]
            batch_data.extend(current_batch)

            bat_file_path = os.path.join(output_folder, f"BAT{batch_count}.bat")
            with open(bat_file_path, "w") as file:
                link = f'start "" https://task.haydigiy.com/admin/order/printorder/?orderId={current_batch[0]}&isPdf=False\n'
                file.write(link)
                file.write('timeout -t 1\n')  # Add the timeout line

                for value in current_batch[1:]:
                    link = f"https://task.haydigiy.com/admin/order/printorder/?orderId={value}&isPdf=False"
                    file.write(f'start "" {link}\n')

            batch_data = []
            remaining_data = remaining_data[batch_size:]
            batch_count += 1

    # Klasör oluştur
    output_folder = "İç Giyim"
    os.makedirs(output_folder, exist_ok=True)

    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    unique_ids_sheet = wb["Unique Ids"]

    # "Id" sütunundaki verileri al
    id_column = unique_ids_sheet["A"][1:]

    # Verileri bir listeye dönüştür
    id_values = [cell.value for cell in id_column if cell.value is not None]

    # .bat dosyalarını oluştur ve klasöre kaydet
    create_bat_files(id_values, output_folder)

    # Excel dosyasını klasöre taşı
    shutil.copy(file_path, os.path.join(output_folder, "İç Giyim.xlsx"))

    gc.collect()

    # Klasör dışında kalan Excel dosyasını sil
    os.remove(file_path)

   


































    # Klasör adları
    folders = ["İç Giyim", "14"]

    # Bugünkü tarihi al
    current_date = datetime.datetime.now().strftime("%Y-%m-%d")

    # Oluşturulacak zip dosyasının adı
    zip_filename = "2500 TL Üzeri Siparişler.zip"

    # Klasörleri kontrol et ve gerektiğinde sil veya zip'e ekle
    with zipfile.ZipFile(zip_filename, 'w') as zipf:
        for folder in folders:
            folder_path = os.path.join(".", folder)
            folder_contents = os.listdir(folder_path)
            bat_files = [file for file in folder_contents if file.endswith(".bat")]

            if bat_files:
                for root, _, files in os.walk(folder_path):
                    for file in files:
                        file_path = os.path.join(root, file)
                        zipf.write(file_path, os.path.relpath(file_path, "."))
                        
                        
            else:
                for root, dirs, files in os.walk(folder_path, topdown=False):
                    for file in files:
                        file_path = os.path.join(root, file)
                        os.remove(file_path)
                    for dir in dirs:
                        dir_path = os.path.join(root, dir)
                        os.rmdir(dir_path)
                os.rmdir(folder_path)





    # Klasörleri sil
    for folder in folders:
        folder_path = os.path.join(".", folder)
        if os.path.exists(folder_path):
            shutil.rmtree(folder_path)





    print(Fore.YELLOW + "2500 TL Üzeri Siparişler Hazırlandı")
    print(" ")
    print(Fore.RED + "Siparişler Entegrasyona Gönderiliyor Ekran Kapanana Kadar İşlem Yapmayın !")

















    #KARGO ENTEGRASYONUNA GÖNDERME

    # Excel dosyasının adını belirtin
    excel_dosyasi = "Kargo Entegrasyonu.xlsx"

    # Excel dosyasını yükle
    df = pd.read_excel(excel_dosyasi)

    # "KargoTakipNumarasi" sütununda dolu olan satırları sil
    df = df[df['KargoTakipNumarasi'].isna()]


    # "KargoFirmasi" sütununda "MNG KARGO" değerine sahip olan satırları sil
    df = df[df['KargoFirmasi'] != 'MNG KARGO']

    # "KargoFirmasi" sütununda "MNG KARGO" değerine sahip olan satırları sil
    df = df[df['KargoFirmasi'] != 'KARGOİST']

    # "Id" sütunu hariç diğer tüm sütunları sil
    df = df[['Id']]

    # Aynı Excel dosyasına kaydet (üzerine yaz)
    df.to_excel(excel_dosyasi, index=False)












    # Excel dosyasını oku
    df = pd.read_excel(excel_dosyasi)

    # "Id" sütunu hariç diğer tüm sütunları sil
    df = df[['Id']].drop_duplicates()

    # Güncellenmiş veriyi aynı Excel dosyasına kaydet (mevcut dosyanın üzerine yazacak)
    df.to_excel(excel_dosyasi, index=False)






    
    # Excel dosyasının adını belirtin
    excel_dosyasi = "Kargo Entegrasyonu.xlsx"


    # Excel dosyasını oku
    df = pd.read_excel(excel_dosyasi)

    # "Id" sütunundaki her verinin başına URL'yi ekleyerek istek gönder
    def send_request(order_id):
        url = f"https://task.haydigiy.com/admin/order/sendordertoshipmentintegration/?orderId={order_id}"

        # Kullanıcı adı ve şifre
        username = "mustafa_kod@haydigiy.com"
        password = "123456"

        # Oturum açılacak web sitesi
        login_url = "https://task.haydigiy.com/kullanici-giris/?ReturnUrl=%2Fadmin"

        # İstek başlıkları
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36",
            "Referer": "https://task.haydigiy.com/",
        }

        # Oturum açma sayfasına GET isteği gönderme
        session = requests.Session()
        response = session.get(login_url, headers=headers)
        soup = BeautifulSoup(response.text, "html.parser")

        # __RequestVerificationToken değerini alma
        token = soup.find("input", {"name": "__RequestVerificationToken"}).get("value")

        # POST isteği için istek verilerini ayarlama
        login_data = {
            "EmailOrPhone": username,
            "Password": password,
            "__RequestVerificationToken": token,
        }

        # Oturum açma isteği gönderme
        response = session.post(login_url, data=login_data, headers=headers)

        # URL'ye istek gönderme
        response = session.get(url, headers=headers)



    # ThreadPoolExecutor kullanarak istekleri paralel olarak gönderme
    with ThreadPoolExecutor(max_workers=5) as executor, tqdm(total=len(df['Id']), desc="Entegrasyona Gönderiliyor") as pbar:
        futures = [executor.submit(send_request, order_id) for order_id in df['Id']]
        for future in futures:
            future.result()  # İşlem tamamlandığında bir sonraki adıma geç
            pbar.update(1)  # İlerleme çubuğunu güncelle




    # Excel dosyasının adı
    excel_dosyasi = "Kargo Entegrasyonu.xlsx"

    # Excel dosyasını sil
    try:
        os.remove(excel_dosyasi)
        pass
    except FileNotFoundError:
        print(f"{excel_dosyasi} adlı Excel dosyası bulunamadı.")
    except Exception as e:
        print(f"Hata oluştu: {e}")

















else:
    pass
        

#endregion

#region Gönderilmesi Gereken Siparişler



if gonderilmesi_gereken_siparisler == "E":

    async def download_file(session, url, index):
        async with session.get(url) as response:
            if response.status == 200:
                file_name = f"link_{index}.xlsx"
                content = await response.read()
                with open(file_name, "wb") as file:
                    file.write(content)
                
            else:
                print(f"Hata! {url} dosyası indirilemedi.")

    async def main():
        base_url = "https://task.haydigiy.com/FaprikaOrderXls/DT6HNF/"
        urls = [f"{base_url}{i}/" for i in range(1, 2)]

        async with aiohttp.ClientSession() as session:
            tasks = [download_file(session, url, index + 1) for index, url in enumerate(urls)]
            await asyncio.gather(*tasks)

        # Excel dosyalarını birleştirip verileri çıkartma
        data_frames = []
        for i in range(1, 2):
            file_name = f"link_{i}.xlsx"
            df = pd.read_excel(file_name)
            data_frames.append(df)

        combined_df = pd.concat(data_frames, ignore_index=True)

        # İstenen sütunları seçme
        selected_columns = ["Id", "OdemeTipi", "TeslimatTelefon", "Barkod", "Adet", "TeslimatEPostaAdresi", "SiparisToplam", "Varyant", "UrunAdi"]
        final_df = combined_df[selected_columns]
        
        # Virgül ve sonrasını kaldırarak stringi sayıya çevirme işlemi
        def clean_and_convert(value):
            try:
                cleaned_value = value.split(',')[0]  # Virgülden önceki kısmı al
                numeric_value = float(cleaned_value)  # Düzenlenmiş veriyi sayıya çevir
                return numeric_value
            except ValueError:
                return None  # Dönüşüm başarısızsa veya veri boşsa None döndür

        # "Adet" ve "SiparisToplam" sütunlarını temizleme ve dönüştürme
        final_df['Adet'] = final_df['Adet'].apply(clean_and_convert)
        final_df['SiparisToplam'] = final_df['SiparisToplam'].apply(clean_and_convert)

        # Düzenlenmiş verileri mevcut dosyanın üzerine kaydetme
        final_df.to_excel("birlesik_excel.xlsx", index=False)
        

        # Birleştirilmiş verileri yeni bir Excel dosyasına kaydetme
        yeni_dosya_adi = "birlesik_excel.xlsx"
        final_df.to_excel(yeni_dosya_adi, index=False)
        

        # İndirilen dosyaları silme
        for i in range(1, 11):
            file_name = f"link_{i}.xlsx"
            if os.path.exists(file_name):
                os.remove(file_name)
                

    if __name__ == "__main__":
        asyncio.run(main())





    
        







    google_sheet_url = "https://docs.google.com/spreadsheets/d/1iKtTb2X28bYDarV3y8YUiRWCnov8zgJBxRVuQOygrNQ/gviz/tq?tqx=out:csv"

    try:
        google_df = pd.read_csv(google_sheet_url)
        google_excel_file = "Hariç Tutulacak Sipariş Numaraları.xlsx"
        google_df.to_excel(google_excel_file, index=False)
    except requests.exceptions.RequestException as e:
        pass



    def main():
        while True:      
            
            # İki Excel dosyasını okuyun
            birlesik_excel = pd.read_excel("birlesik_excel.xlsx")
            haric_excel = pd.read_excel("Hariç Tutulacak Sipariş Numaraları.xlsx")

            # İlk sütunlara göre birleşik Excel verisinden, haric Excel verisinde bulunan satırları filtreleyin
            birlesik_excel = birlesik_excel[~birlesik_excel['Id'].isin(haric_excel['Id'])]

            # Sonucu mevcut "birlesik_excel.xlsx" dosyasına kaydedin (var olan dosyanın üstüne yazacak)
            birlesik_excel.to_excel("birlesik_excel.xlsx", index=False)
            break
    if __name__ == "__main__":
        main()


    # Kaynak dosya adı
    kaynak_excel = "birlesik_excel.xlsx"

    # Kopya dosya adı (istediğiniz adı ve konumu belirtin)
    kopya_excel = "Hazırlanan Sipariş Numaraları2.xlsx"

    # Dosyayı kopyala
    shutil.copy(kaynak_excel, kopya_excel)




    file_path = "Hariç Tutulacak Sipariş Numaraları.xlsx"

    if os.path.exists(file_path):
        os.remove(file_path)















































    # Excel dosyasını okuma
    df = pd.read_excel("birlesik_excel.xlsx")


    # İşlemi gerçekleştiren fonksiyon
    def duplicate_rows(row):
        count = int(row["Adet"])
        return pd.concat([row] * count, axis=1).T

    # Tüm satırları işleme tabi tutma
    new_rows = df.apply(duplicate_rows, axis=1)

    # Yeni veri çerçevesini oluşturma
    new_df = pd.concat(new_rows.tolist(), ignore_index=True)

    # Sadece belirtilen sütunları seçme
    selected_columns = ["Id", "Barkod", "UrunAdi", "Varyant"]
    new_df = new_df[selected_columns]

    # Veriyi yeni bir Excel dosyasına yazma
    new_df.to_excel("birlesik_excel.xlsx", index=False)

    











    url = "https://haydigiy.online/Products/rafkodlari.php"
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")
    table = soup.find("table")
    data = []
    for row in table.find_all("tr"):
        row_data = []
        for cell in row.find_all(["th", "td"]):
            row_data.append(cell.get_text(strip=True))
        data.append(row_data)
    df = pd.DataFrame(data[1:], columns=data[0])
    df.to_excel("Raf Kodu.xlsx", index=False)
        
        






    # "birlesik_excel.xlsx" ve "Raf Kodu.xlsx" dosyalarını okuma
    sonuc_df = pd.read_excel("birlesik_excel.xlsx")
    google_sheet_df = pd.read_excel("Raf Kodu.xlsx")

    # "birlesik_excel.xlsx" dosyasına yeni bir sütun ekleyerek başlangıçta "Raf Kodu Yok" değerleri ile doldurma
    sonuc_df["GoogleSheetVerisi"] = "Raf Kodu Yok"

    # Her bir "Barkod" değeri için işlem yapma
    for index, row in sonuc_df.iterrows():
        barkod = row["Barkod"]
        
        # "Raf Kodu.xlsx" dosyasında ilgili "Barkod"u arama
        matching_row = google_sheet_df[google_sheet_df.iloc[:, 0] == barkod]
        
        # Eşleşen "Barkod" varsa ve karşılık gelen hücre boş değilse, değeri "GoogleSheetVerisi" sütununa yazma
        if not matching_row.empty and not pd.isnull(matching_row.iloc[0, 2]):
            sonuc_df.at[index, "GoogleSheetVerisi"] = matching_row.iloc[0, 2]

    # "birlesik_excel.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("birlesik_excel.xlsx", index=False)
  






    # "birlesik_excel.xlsx" dosyasını güncelleme
    sonuc_df["GoogleSheetVerisi Kopya"] = sonuc_df["GoogleSheetVerisi"]  # "GoogleSheetVerisi" sütununu kopyala
    sonuc_df["GoogleSheetVerisi Kopya"] = sonuc_df["GoogleSheetVerisi Kopya"].str.split("-", n=1).str[0]  # "-" den sonrasını temizle
    sonuc_df["GoogleSheetVerisi Kopya"] = pd.to_numeric(sonuc_df["GoogleSheetVerisi Kopya"], errors="coerce")  # Sayıya dönüştür
    sonuc_df = sonuc_df.sort_values(by="GoogleSheetVerisi Kopya")  # "GoogleSheetVerisi Kopya" sütununa göre sırala


    sonuc_df.to_excel("birlesik_excel.xlsx", index=False)
    













    # "birlesik_excel.xlsx" ve "Raf Kodu.xlsx" dosyalarını okuma
    sonuc_df = pd.read_excel("birlesik_excel.xlsx")
    google_sheet_df = pd.read_excel("Raf Kodu.xlsx")

    # "birlesik_excel.xlsx" dosyasına yeni bir sütun ekleyerek başlangıçta "Raf Kodu Yok" değerleri ile doldurma
    sonuc_df["Kategori"] = "Raf Kodu Yok"

    # Her bir "Barkod" değeri için işlem yapma
    for index, row in sonuc_df.iterrows():
        barkod = row["Barkod"]
        
        # "Raf Kodu.xlsx" dosyasında ilgili "Barkod"u arama
        matching_row = google_sheet_df[google_sheet_df.iloc[:, 0] == barkod]
        
        # Eşleşen "Barkod" varsa ve karşılık gelen hücre boş değilse, değeri "GoogleSheetVerisi" sütununa yazma
        if not matching_row.empty and not pd.isnull(matching_row.iloc[0, 3]):
            sonuc_df.at[index, "Kategori"] = matching_row.iloc[0, 3]

    # "birlesik_excel.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("birlesik_excel.xlsx", index=False)
   





    # Excel dosyasını oku
    excel_file_path = "birlesik_excel.xlsx"  # Excel dosyasının adını ve yolunu belirtin
    df = pd.read_excel(excel_file_path)

    # Yeni bir sütun ekleyerek işlem sonuçlarını tut
    df["Yeni Kategori"] = ""

    # İç Giyim içeren satırları işle
    innerwear_rows = df[df["Kategori"].str.contains("İç Giyim")]

    # İç Giyim içeren satırları işle
    for index, row in innerwear_rows.iterrows():
        df.loc[index, "Yeni Kategori"] = "İç Giyim"

    # Sonucu kaydet
    output_file_path = "birlesik_excel.xlsx"  # Sonucun kaydedileceği dosya adı ve yolunu belirtin
    df.to_excel(output_file_path, index=False)

    

    # Excel dosyasını oku
    excel_file_path = "birlesik_excel.xlsx"  # Excel dosyasının adını ve yolunu belirtin
    df = pd.read_excel(excel_file_path)

    # "Id" değerlerine göre grupla ve işlemi yap
    grouped = df.groupby("Id")

    # "Id" değerlerine göre "Yeni Kategori" değerini güncelle
    for group_name, group_data in grouped:
        if any(row["Yeni Kategori"] == "İç Giyim" for _, row in group_data.iterrows()):
            df.loc[df["Id"] == group_name, "Yeni Kategori"] = "İç Giyim"

    # Sonucu kaydet
    output_file_path = "birlesik_excel.xlsx"  # Sonucun kaydedileceği dosya adı ve yolunu belirtin
    df.to_excel(output_file_path, index=False)

   







    # Excel dosyasını oku
    excel_file_path = "birlesik_excel.xlsx"  # Ana Excel dosyasının adını ve yolunu belirtin
    df = pd.read_excel(excel_file_path)

    # "Yeni Kategori" değeri "İç Giyim" olan satırları seç
    innerwear_rows = df[df["Yeni Kategori"] == "İç Giyim"]

    # Ayrı Excel dosyasına kaydet
    output_file_path = "İç Giyim.xlsx"  # Ayrı dosyanın adını ve yolunu belirtin
    innerwear_rows.to_excel(output_file_path, index=False)

    # Ana DataFrame'den "İç Giyim" satırları sil
    df = df[df["Yeni Kategori"] != "İç Giyim"]
    df.drop(columns=["Yeni Kategori"], inplace=True)  # "Yeni Kategori" sütununu sil

    # Ana Excel dosyasını güncelle
    df.to_excel(excel_file_path, index=False)

   












    excel_to_delete = "Raf Kodu.xlsx"

    # Excel dosyasını sil
    if os.path.exists(excel_to_delete):
        os.remove(excel_to_delete)
        
    else:
        print("Dosya bulunamadı:", excel_to_delete)



















    old_file_path = "birlesik_excel.xlsx"
    new_file_path = "Kalanlar.xlsx"

    # Dosyanın adını değiştir
    os.rename(old_file_path, new_file_path)

    







    # Excel dosyasını oku
    excel_file_path = "Kalanlar.xlsx"  # Kalanlar Excel dosyasının adını ve yolunu belirtin
    df = pd.read_excel(excel_file_path)

    # "GoogleSheetVerisi Kopya" ve "Kategori" sütunlarını sil
    columns_to_drop = ["GoogleSheetVerisi Kopya", "Kategori"]
    df.drop(columns=columns_to_drop, inplace=True)

    # Dosyayı güncelle
    df.to_excel(excel_file_path, index=False)

    





    # Excel dosyasını oku
    excel_file_path = "İç Giyim.xlsx"  # Kalanlar Excel dosyasının adını ve yolunu belirtin
    df = pd.read_excel(excel_file_path)

    # "GoogleSheetVerisi Kopya" ve "Kategori" sütunlarını sil
    columns_to_drop = ["GoogleSheetVerisi Kopya", "Kategori", "Yeni Kategori"]
    df.drop(columns=columns_to_drop, inplace=True)

    # Dosyayı güncelle
    df.to_excel(excel_file_path, index=False)

    


























































    sonuc_df = pd.read_excel("Kalanlar.xlsx")

    # "Kalanlar.xlsx" dosyasını güncelleme
    sonuc_df["UrunAdi Kopya"] = sonuc_df["UrunAdi"]  # "UrunAdi" sütununu kopyala
    sonuc_df["UrunAdi Kopya"] = sonuc_df["UrunAdi Kopya"].str.split("-", n=1).str[1]  # "-" den öncesini temizle

    # "UrunAdi Kopya" sütununu "Kalanlar.xlsx" dosyasına ekleyerek güncelleme
    with pd.ExcelWriter("Kalanlar.xlsx") as writer:
        sonuc_df.to_excel(writer, index=False)

    








    # "UrunAdi" sütununu en sağına yapıştırma
    sonuc_df["UrunAdiKopya2"] = sonuc_df["UrunAdi"]

    # "UrunAdiKopya2" sütununda " - " dan sonrasını ve son boşluktan öncesini silme
    sonuc_df["UrunAdiKopya2"] = sonuc_df["UrunAdiKopya2"].apply(lambda x: x.split(" - ")[0].rsplit(" ", 1)[-1].strip() if " - " in x else x)

    # "UrunAdiKopya2" sütununu en sağa taşıma
    column_order = list(sonuc_df.columns)
    column_order.remove("UrunAdiKopya2")
    column_order.append("UrunAdiKopya2")
    sonuc_df = sonuc_df[column_order]

    # "Kalanlar.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("Kalanlar.xlsx", index=False)
    






    # "UrunAdi" sütununu tablonun sonuna bir kez daha kopyalama ve düzenleme
    sonuc_df["UrunAdiKopya3"] = sonuc_df["UrunAdi"].apply(lambda x: x.split(" - ", 1)[0].strip() if " - " in x else x)

    # "UrunAdiKopya3" sütununu en sağa taşıma
    column_order = list(sonuc_df.columns)
    column_order.remove("UrunAdiKopya3")
    column_order.append("UrunAdiKopya3")
    sonuc_df = sonuc_df[column_order]

    # "Kalanlar.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("Kalanlar.xlsx", index=False)
    









    # Verileri birleştirip yeni sütun oluşturma
    sonuc_df["BirlesikVeri"] = sonuc_df["UrunAdi Kopya"] + " - " + sonuc_df["UrunAdiKopya2"] + " - " + sonuc_df["Varyant"]

    # "BirlesikVeri" sütununu en sağa taşıma
    column_order = list(sonuc_df.columns)
    column_order.remove("BirlesikVeri")
    column_order.append("BirlesikVeri")
    sonuc_df = sonuc_df[column_order]

    # "Kalanlar.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("Kalanlar.xlsx", index=False)
    

    # "BirlesikVeri" sütunundaki "Beden:" ibaresini çıkarma
    sonuc_df["BirlesikVeri"] = sonuc_df["BirlesikVeri"].str.replace("Beden:", "")

    # "Kalanlar.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("Kalanlar.xlsx", index=False)
    





    # Belirtilen sütunları silme
    columns_to_drop = ["UrunAdi", "Varyant", "UrunAdi Kopya", "UrunAdiKopya2"]
    sonuc_df.drop(columns_to_drop, axis=1, inplace=True)

    # "Kalanlar.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("Kalanlar.xlsx", index=False)
    






    # "Id" sütununu teke düşürme
    unique_ids = sonuc_df["Id"].drop_duplicates()

    # Yeni bir Excel sayfası oluşturma ve "Id" değerlerini yazma
    with pd.ExcelWriter("Kalanlar.xlsx", engine="openpyxl", mode="a") as writer:
        unique_ids.to_excel(writer, sheet_name="Unique Ids", index=False)

    










    

    # Sonuç dosyasını yükle
    file_path = "Kalanlar.xlsx"
    wb = load_workbook(file_path)
    sheet = wb["Unique Ids"]

    # Başlangıç sütunu ve satırı
    start_column = 2
    start_row = 2

    # Toplam tekrar sayısı ve her tekrardaki numara adedi
    repeat_count = 100
    numbers_per_repeat = 14

    # Verileri ekleme
    for _ in range(repeat_count):
        for num in range(1, numbers_per_repeat + 1):
            sheet.cell(row=start_row, column=start_column, value=num)
            start_row += 1

    # Değişiklikleri kaydetme
    wb.save("Kalanlar.xlsx")






    # Sonuç dosyasını yükle
    file_path = "Kalanlar.xlsx"
    wb = load_workbook(file_path)
    sheet = wb["Unique Ids"]

    # Başlangıç sütunu ve satırı
    start_column = 3
    start_row = 2

    # Toplam tekrar sayısı ve her tekrardaki numara adedi
    repeat_count = 14
    numbers_per_repeat = 100

    # Verileri ekleme
    for num in range(1, numbers_per_repeat + 1):
        for repeat in range(repeat_count):
            sheet.cell(row=start_row, column=start_column, value=num)
            start_row += 1

    # Değişiklikleri kaydetme
    wb.save("Kalanlar.xlsx")

    









    # Sonuç dosyasını yükle
    file_path = "Kalanlar.xlsx"
    wb = load_workbook(file_path)
    unique_ids_sheet = wb["Unique Ids"]
    main_sheet = wb["Sheet1"]

    # "Id" sütununun verilerini al
    id_column = main_sheet["A"][1:]
    unique_ids_column = unique_ids_sheet["A"][1:]

    # Karşılık gelen değerleri bulup "Sheet1" sayfasının en sağında yeni bir sütuna ekle
    new_column = main_sheet.max_column + 1
    main_sheet.cell(row=1, column=new_column, value="Matching Value")

    for id_cell in id_column:
        id_value = id_cell.value
        for unique_id_cell in unique_ids_column:
            if unique_id_cell.value == id_value:
                matching_value = unique_id_cell.offset(column=1).value
                main_sheet.cell(row=id_cell.row, column=new_column, value=matching_value)
                break

    # Değişiklikleri kaydetme
    wb.save("Kalanlar.xlsx")

    










    # Sonuç dosyasını yükle
    file_path = "Kalanlar.xlsx"
    wb = load_workbook(file_path)
    unique_ids_sheet = wb["Unique Ids"]
    main_sheet = wb["Sheet1"]

    # "Id" sütununun verilerini al
    id_column = main_sheet["A"][1:]
    unique_ids_column = unique_ids_sheet["A"][1:]

    # Karşılık gelen değerleri bulup "Sheet1" sayfasının en sağında yeni bir sütuna ekle
    new_column = main_sheet.max_column + 1
    main_sheet.cell(row=1, column=new_column, value="Matching Value (3rd Column)")

    for id_cell in id_column:
        id_value = id_cell.value
        for unique_id_cell in unique_ids_column:
            if unique_id_cell.value == id_value:
                matching_value = unique_id_cell.offset(column=2).value
                main_sheet.cell(row=id_cell.row, column=new_column, value=matching_value)
                break

    # Değişiklikleri kaydetme
    wb.save("Kalanlar.xlsx")

   









    # Sonuç dosyasını yükle
    file_path = "Kalanlar.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Sütun başlıklarını değiştir
    new_column_titles = {
        "Id": "SiparişNO",
        "BirlesikVeri": "ÜRÜN",
        "GoogleSheetVerisi": "RAF KODU",
        "UrunAdiKopya3": "ÜRÜN ADI",
        "Matching Value": "KUTU",
        "Matching Value (3rd Column)": "ÇN"
    }

    for col_idx, col_name in enumerate(main_sheet.iter_cols(min_col=1, max_col=main_sheet.max_column), start=1):
        old_title = col_name[0].value
        new_title = new_column_titles.get(old_title, old_title)
        col_name[0].value = new_title

    # Değişiklikleri kaydetme
    wb.save("Kalanlar.xlsx")

    








    

    # Sonuç dosyasını yükle
    file_path = "Kalanlar.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Sütunların yeni sıralaması
    new_column_order = [
        "RAF KODU",
        "ÜRÜN",
        "Barkod",
        "KUTU",
        "ÜRÜN ADI",
        "ÇN",
        "SiparişNO"
    ]

    # Yeni bir DataFrame oluştur
    data = main_sheet.iter_rows(min_row=2, values_only=True)
    df = pd.DataFrame(data, columns=[cell.value for cell in main_sheet[1]])

    # Sütunları yeni sıralamaya göre düzenle
    df = df[new_column_order]

    # Mevcut başlıkları güncelle
    for idx, column_name in enumerate(new_column_order, start=1):
        main_sheet.cell(row=1, column=idx, value=column_name)

    # DataFrame verilerini sayfaya yaz
    for r_idx, row in enumerate(df.values, 2):
        for c_idx, value in enumerate(row, 1):
            main_sheet.cell(row=r_idx, column=c_idx, value=value)

    # Değişiklikleri kaydet
    wb.save("Kalanlar.xlsx")

    







    

    # Sonuç dosyasını yükle
    file_path = "Kalanlar.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Hücreleri ortala ve ortaya hizala
    for row in main_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")


    # Değişiklikleri kaydet
    wb.save("Kalanlar.xlsx")

    






    
    

    # Sonuç dosyasını yükle
    file_path = "Kalanlar.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Kenarlık stili oluştur
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Hücreleri kalın yap, yazı fontunu 14 yap ve kenarlık ekle
    for row in main_sheet.iter_rows():
        for cell in row:
            cell.font = Font(bold=True, size=14)
            cell.border = border_style

    # Değişiklikleri kaydet
    wb.save("Kalanlar.xlsx")

    


    

    # Sonuç dosyasını yükle
    file_path = "Kalanlar.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # "RAF KODU" sütununu 45 piksel yap
    main_sheet.column_dimensions["C"].width = 45

    # Tüm hücreleri en uygun sütun genişliği olarak ayarla
    for column in main_sheet.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        main_sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Değişiklikleri kaydet
    wb.save("Kalanlar.xlsx")

    



    # Sonuç dosyasını yükle
    file_path = "Kalanlar.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # İlk sütunu (A sütunu) 45 piksel genişliğinde yap
    main_sheet.column_dimensions["A"].width = 45
    main_sheet.column_dimensions["C"].width = 14
    main_sheet.column_dimensions["G"].width = 14
    main_sheet.column_dimensions["D"].width = 9
    main_sheet.column_dimensions["F"].width = 5

    # Değişiklikleri kaydet
    wb.save("Kalanlar.xlsx")

   








    

    # Sonuç dosyasını yükle
    file_path = "Kalanlar.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Tüm hücrelere "Metni Kaydır" formatını uygula
    for row in main_sheet.iter_rows():
        for cell in row:
            new_alignment = copy(cell.alignment)
            new_alignment.wrap_text = True
            cell.alignment = new_alignment

    # Değişiklikleri kaydet
    wb.save("Kalanlar.xlsx")

   







    

    # Sonuç dosyasını yükle
    file_path = "Kalanlar.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Tabloyu oluşturma
    table = Table(displayName="MyTable", ref=main_sheet.dimensions)

    # Tablo stili oluşturma (gri-beyaz)
    style = TableStyleInfo(
        name="TableStyleLight1", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True
    )

    # Tabloya stil atama
    table.tableStyleInfo = style

    # Tabloyu sayfaya ekleme
    main_sheet.add_table(table)

    # Değişiklikleri kaydetme
    wb.save("Kalanlar.xlsx")

    


    

    def create_bat_files(data, output_folder, batch_size=14):
        batch_count = 1
        batch_data = []
        remaining_data = data

        while len(remaining_data) > 0:
            current_batch = remaining_data[:batch_size]
            batch_data.extend(current_batch)

            bat_file_path = os.path.join(output_folder, f"BAT{batch_count}.bat")
            with open(bat_file_path, "w") as file:
                link = f'start "" https://task.haydigiy.com/admin/order/printorder/?orderId={current_batch[0]}&isPdf=False\n'
                file.write(link)
                file.write('timeout -t 1\n')  # Add the timeout line

                for value in current_batch[1:]:
                    link = f"https://task.haydigiy.com/admin/order/printorder/?orderId={value}&isPdf=False"
                    file.write(f'start "" {link}\n')

            batch_data = []
            remaining_data = remaining_data[batch_size:]
            batch_count += 1

    # Klasör oluştur
    output_folder = "Kalanlar"
    os.makedirs(output_folder, exist_ok=True)

    # Sonuç dosyasını yükle
    file_path = "Kalanlar.xlsx"
    wb = load_workbook(file_path)
    unique_ids_sheet = wb["Unique Ids"]

    # "Id" sütunundaki verileri al
    id_column = unique_ids_sheet["A"][1:]

    # Verileri bir listeye dönüştür
    id_values = [cell.value for cell in id_column if cell.value is not None]

    # .bat dosyalarını oluştur ve klasöre kaydet
    create_bat_files(id_values, output_folder)

    # Excel dosyasını klasöre taşı
    shutil.copy(file_path, os.path.join(output_folder, "Kalanlar.xlsx"))

    gc.collect()

    # Klasör dışında kalan Excel dosyasını sil
    os.remove(file_path)

    









    sonuc_df = pd.read_excel("İç Giyim.xlsx")

    # "İç Giyim.xlsx" dosyasını güncelleme
    sonuc_df["UrunAdi Kopya"] = sonuc_df["UrunAdi"]  # "UrunAdi" sütununu kopyala
    sonuc_df["UrunAdi Kopya"] = sonuc_df["UrunAdi Kopya"].str.split("-", n=1).str[1]  # "-" den öncesini temizle

    # "UrunAdi Kopya" sütununu "İç Giyim.xlsx" dosyasına ekleyerek güncelleme
    with pd.ExcelWriter("İç Giyim.xlsx") as writer:
        sonuc_df.to_excel(writer, index=False)

   








    # "UrunAdi" sütununu en sağına yapıştırma
    sonuc_df["UrunAdiKopya2"] = sonuc_df["UrunAdi"]

    # "UrunAdiKopya2" sütununda " - " dan sonrasını ve son boşluktan öncesini silme
    sonuc_df["UrunAdiKopya2"] = sonuc_df["UrunAdiKopya2"].apply(lambda x: x.split(" - ")[0].rsplit(" ", 1)[-1].strip() if " - " in x else x)

    # "UrunAdiKopya2" sütununu en sağa taşıma
    column_order = list(sonuc_df.columns)
    column_order.remove("UrunAdiKopya2")
    column_order.append("UrunAdiKopya2")
    sonuc_df = sonuc_df[column_order]

    # "İç Giyim.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("İç Giyim.xlsx", index=False)
   






    # "UrunAdi" sütununu tablonun sonuna bir kez daha kopyalama ve düzenleme
    sonuc_df["UrunAdiKopya3"] = sonuc_df["UrunAdi"].apply(lambda x: x.split(" - ", 1)[0].strip() if " - " in x else x)

    # "UrunAdiKopya3" sütununu en sağa taşıma
    column_order = list(sonuc_df.columns)
    column_order.remove("UrunAdiKopya3")
    column_order.append("UrunAdiKopya3")
    sonuc_df = sonuc_df[column_order]

    # "İç Giyim.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("İç Giyim.xlsx", index=False)









    # Verileri birleştirip yeni sütun oluşturma
    sonuc_df["BirlesikVeri"] = sonuc_df["UrunAdi Kopya"] + " - " + sonuc_df["UrunAdiKopya2"] + " - " + sonuc_df["Varyant"]

    # "BirlesikVeri" sütununu en sağa taşıma
    column_order = list(sonuc_df.columns)
    column_order.remove("BirlesikVeri")
    column_order.append("BirlesikVeri")
    sonuc_df = sonuc_df[column_order]

    # "İç Giyim.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("İç Giyim.xlsx", index=False)
    

    # "BirlesikVeri" sütunundaki "Beden:" ibaresini çıkarma
    sonuc_df["BirlesikVeri"] = sonuc_df["BirlesikVeri"].str.replace("Beden:", "")

    # "İç Giyim.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("İç Giyim.xlsx", index=False)
   





    # Belirtilen sütunları silme
    columns_to_drop = ["UrunAdi", "Varyant", "UrunAdi Kopya", "UrunAdiKopya2"]
    sonuc_df.drop(columns_to_drop, axis=1, inplace=True)

    # "İç Giyim.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("İç Giyim.xlsx", index=False)
 






    # "Id" sütununu teke düşürme
    unique_ids = sonuc_df["Id"].drop_duplicates()

    # Yeni bir Excel sayfası oluşturma ve "Id" değerlerini yazma
    with pd.ExcelWriter("İç Giyim.xlsx", engine="openpyxl", mode="a") as writer:
        unique_ids.to_excel(writer, sheet_name="Unique Ids", index=False)

   










    

    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    sheet = wb["Unique Ids"]

    # Başlangıç sütunu ve satırı
    start_column = 2
    start_row = 2

    # Toplam tekrar sayısı ve her tekrardaki numara adedi
    repeat_count = 100
    numbers_per_repeat = 14

    # Verileri ekleme
    for _ in range(repeat_count):
        for num in range(1, numbers_per_repeat + 1):
            sheet.cell(row=start_row, column=start_column, value=num)
            start_row += 1

    # Değişiklikleri kaydetme
    wb.save("İç Giyim.xlsx")






    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    sheet = wb["Unique Ids"]

    # Başlangıç sütunu ve satırı
    start_column = 3
    start_row = 2

    # Toplam tekrar sayısı ve her tekrardaki numara adedi
    repeat_count = 14
    numbers_per_repeat = 100

    # Verileri ekleme
    for num in range(1, numbers_per_repeat + 1):
        for repeat in range(repeat_count):
            sheet.cell(row=start_row, column=start_column, value=num)
            start_row += 1

    # Değişiklikleri kaydetme
    wb.save("İç Giyim.xlsx")










    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    unique_ids_sheet = wb["Unique Ids"]
    main_sheet = wb["Sheet1"]

    # "Id" sütununun verilerini al
    id_column = main_sheet["A"][1:]
    unique_ids_column = unique_ids_sheet["A"][1:]

    # Karşılık gelen değerleri bulup "Sheet1" sayfasının en sağında yeni bir sütuna ekle
    new_column = main_sheet.max_column + 1
    main_sheet.cell(row=1, column=new_column, value="Matching Value")

    for id_cell in id_column:
        id_value = id_cell.value
        for unique_id_cell in unique_ids_column:
            if unique_id_cell.value == id_value:
                matching_value = unique_id_cell.offset(column=1).value
                main_sheet.cell(row=id_cell.row, column=new_column, value=matching_value)
                break

    # Değişiklikleri kaydetme
    wb.save("İç Giyim.xlsx")

  










    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    unique_ids_sheet = wb["Unique Ids"]
    main_sheet = wb["Sheet1"]

    # "Id" sütununun verilerini al
    id_column = main_sheet["A"][1:]
    unique_ids_column = unique_ids_sheet["A"][1:]

    # Karşılık gelen değerleri bulup "Sheet1" sayfasının en sağında yeni bir sütuna ekle
    new_column = main_sheet.max_column + 1
    main_sheet.cell(row=1, column=new_column, value="Matching Value (3rd Column)")

    for id_cell in id_column:
        id_value = id_cell.value
        for unique_id_cell in unique_ids_column:
            if unique_id_cell.value == id_value:
                matching_value = unique_id_cell.offset(column=2).value
                main_sheet.cell(row=id_cell.row, column=new_column, value=matching_value)
                break

    # Değişiklikleri kaydetme
    wb.save("İç Giyim.xlsx")

    









    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Sütun başlıklarını değiştir
    new_column_titles = {
        "Id": "SiparişNO",
        "BirlesikVeri": "ÜRÜN",
        "GoogleSheetVerisi": "RAF KODU",
        "UrunAdiKopya3": "ÜRÜN ADI",
        "Matching Value": "KUTU",
        "Matching Value (3rd Column)": "ÇN"
    }

    for col_idx, col_name in enumerate(main_sheet.iter_cols(min_col=1, max_col=main_sheet.max_column), start=1):
        old_title = col_name[0].value
        new_title = new_column_titles.get(old_title, old_title)
        col_name[0].value = new_title

    # Değişiklikleri kaydetme
    wb.save("İç Giyim.xlsx")

   








    

    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Sütunların yeni sıralaması
    new_column_order = [
        "RAF KODU",
        "ÜRÜN",
        "Barkod",
        "KUTU",
        "ÜRÜN ADI",
        "ÇN",
        "SiparişNO"
    ]

    # Yeni bir DataFrame oluştur
    data = main_sheet.iter_rows(min_row=2, values_only=True)
    df = pd.DataFrame(data, columns=[cell.value for cell in main_sheet[1]])

    # Sütunları yeni sıralamaya göre düzenle
    df = df[new_column_order]

    # Mevcut başlıkları güncelle
    for idx, column_name in enumerate(new_column_order, start=1):
        main_sheet.cell(row=1, column=idx, value=column_name)

    # DataFrame verilerini sayfaya yaz
    for r_idx, row in enumerate(df.values, 2):
        for c_idx, value in enumerate(row, 1):
            main_sheet.cell(row=r_idx, column=c_idx, value=value)

    # Değişiklikleri kaydet
    wb.save("İç Giyim.xlsx")

    







    

    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Hücreleri ortala ve ortaya hizala
    for row in main_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")


    # Değişiklikleri kaydet
    wb.save("İç Giyim.xlsx")

    






    
    

    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Kenarlık stili oluştur
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Hücreleri kalın yap, yazı fontunu 14 yap ve kenarlık ekle
    for row in main_sheet.iter_rows():
        for cell in row:
            cell.font = Font(bold=True, size=14)
            cell.border = border_style

    # Değişiklikleri kaydet
    wb.save("İç Giyim.xlsx")

    


    

    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # "RAF KODU" sütununu 45 piksel yap
    main_sheet.column_dimensions["C"].width = 45

    # Tüm hücreleri en uygun sütun genişliği olarak ayarla
    for column in main_sheet.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        main_sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Değişiklikleri kaydet
    wb.save("İç Giyim.xlsx")

    



    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # İlk sütunu (A sütunu) 45 piksel genişliğinde yap
    main_sheet.column_dimensions["A"].width = 45
    main_sheet.column_dimensions["C"].width = 14
    main_sheet.column_dimensions["G"].width = 14
    main_sheet.column_dimensions["D"].width = 9
    main_sheet.column_dimensions["F"].width = 5

    # Değişiklikleri kaydet
    wb.save("İç Giyim.xlsx")

    








    

    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Tüm hücrelere "Metni Kaydır" formatını uygula
    for row in main_sheet.iter_rows():
        for cell in row:
            new_alignment = copy(cell.alignment)
            new_alignment.wrap_text = True
            cell.alignment = new_alignment

    # Değişiklikleri kaydet
    wb.save("İç Giyim.xlsx")

    







    

    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Tabloyu oluşturma
    table = Table(displayName="MyTable", ref=main_sheet.dimensions)

    # Tablo stili oluşturma (gri-beyaz)
    style = TableStyleInfo(
        name="TableStyleLight1", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True
    )

    # Tabloya stil atama
    table.tableStyleInfo = style

    # Tabloyu sayfaya ekleme
    main_sheet.add_table(table)

    # Değişiklikleri kaydetme
    wb.save("İç Giyim.xlsx")

   


    

    def create_bat_files(data, output_folder, batch_size=14):
        batch_count = 1
        batch_data = []
        remaining_data = data

        while len(remaining_data) > 0:
            current_batch = remaining_data[:batch_size]
            batch_data.extend(current_batch)

            bat_file_path = os.path.join(output_folder, f"BAT{batch_count}.bat")
            with open(bat_file_path, "w") as file:
                link = f'start "" https://task.haydigiy.com/admin/order/printorder/?orderId={current_batch[0]}&isPdf=False\n'
                file.write(link)
                file.write('timeout -t 1\n')  # Add the timeout line

                for value in current_batch[1:]:
                    link = f"https://task.haydigiy.com/admin/order/printorder/?orderId={value}&isPdf=False"
                    file.write(f'start "" {link}\n')

            batch_data = []
            remaining_data = remaining_data[batch_size:]
            batch_count += 1

    # Klasör oluştur
    output_folder = "İç Giyim"
    os.makedirs(output_folder, exist_ok=True)

    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    unique_ids_sheet = wb["Unique Ids"]

    # "Id" sütunundaki verileri al
    id_column = unique_ids_sheet["A"][1:]

    # Verileri bir listeye dönüştür
    id_values = [cell.value for cell in id_column if cell.value is not None]

    # .bat dosyalarını oluştur ve klasöre kaydet
    create_bat_files(id_values, output_folder)

    # Excel dosyasını klasöre taşı
    shutil.copy(file_path, os.path.join(output_folder, "İç Giyim.xlsx"))

    gc.collect()

    # Klasör dışında kalan Excel dosyasını sil
    os.remove(file_path)

    

































    
    
    

    # Klasör adları
    folders = ["İç Giyim", "Kalanlar"]

    # Bugünkü tarihi al
    current_date = datetime.datetime.now().strftime("%Y-%m-%d")

    # Oluşturulacak zip dosyasının adı
    zip_filename = "Gönderilmesi Gereken Siparişler.zip"

    # Klasörleri kontrol et ve gerektiğinde sil veya zip'e ekle
    with zipfile.ZipFile(zip_filename, 'w') as zipf:
        for folder in folders:
            folder_path = os.path.join(".", folder)
            folder_contents = os.listdir(folder_path)
            bat_files = [file for file in folder_contents if file.endswith(".bat")]

            if bat_files:
                for root, _, files in os.walk(folder_path):
                    for file in files:
                        file_path = os.path.join(root, file)
                        zipf.write(file_path, os.path.relpath(file_path, "."))
                
                
            else:
                for root, dirs, files in os.walk(folder_path, topdown=False):
                    for file in files:
                        file_path = os.path.join(root, file)
                        os.remove(file_path)
                    for dir in dirs:
                        dir_path = os.path.join(root, dir)
                        os.rmdir(dir_path)
                os.rmdir(folder_path)





    # Klasörleri sil
    for folder in folders:
        folder_path = os.path.join(".", folder)
        if os.path.exists(folder_path):
            shutil.rmtree(folder_path)



    print(Fore.YELLOW + "Gönderilmesi Gereken Siparişler Hazırlandı")
else: 
    pass


#endregion

#region Faturasız Siparişler

if faturasiz_siparisler == "E":

    async def download_file(session, url, index):
        async with session.get(url) as response:
            if response.status == 200:
                file_name = f"link_{index}.xlsx"
                content = await response.read()
                with open(file_name, "wb") as file:
                    file.write(content)
                
            else:
                print(f"Hata! {url} dosyası indirilemedi.")

    async def main():
        base_url = "https://task.haydigiy.com/FaprikaOrderXls/DV7JI2/"
        urls = [f"{base_url}{i}/" for i in range(1, 2)]

        async with aiohttp.ClientSession() as session:
            tasks = [download_file(session, url, index + 1) for index, url in enumerate(urls)]
            await asyncio.gather(*tasks)

        # Excel dosyalarını birleştirip verileri çıkartma
        data_frames = []
        for i in range(1, 2):
            file_name = f"link_{i}.xlsx"
            df = pd.read_excel(file_name)
            data_frames.append(df)

        combined_df = pd.concat(data_frames, ignore_index=True)

        # İstenen sütunları seçme
        selected_columns = ["Id", "OdemeTipi", "SiparisDurumu", "TeslimatTelefon", "Barkod", "Adet", "TeslimatEPostaAdresi", "SiparisToplam", "Varyant", "UrunAdi"]
        final_df = combined_df[selected_columns]
        
        # "TeslimatTelefon" sütununda replace yapma
        combined_df['TeslimatTelefon'] = combined_df['TeslimatTelefon'].replace(r'[()/-]', '', regex=True)

        # Virgül ve sonrasını kaldırarak stringi sayıya çevirme işlemi
        def clean_and_convert(value):
            try:
                cleaned_value = value.split(',')[0]  # Virgülden önceki kısmı al
                numeric_value = float(cleaned_value)  # Düzenlenmiş veriyi sayıya çevir
                return numeric_value
            except ValueError:
                return None  # Dönüşüm başarısızsa veya veri boşsa None döndür

        # "Adet" ve "SiparisToplam" sütunlarını temizleme ve dönüştürme
        final_df['Adet'] = final_df['Adet'].apply(clean_and_convert)
        final_df['SiparisToplam'] = final_df['SiparisToplam'].apply(clean_and_convert)

        # Düzenlenmiş verileri mevcut dosyanın üzerine kaydetme
        final_df.to_excel("birlesik_excel.xlsx", index=False)
        

        # Birleştirilmiş verileri yeni bir Excel dosyasına kaydetme
        yeni_dosya_adi = "birlesik_excel.xlsx"
        final_df.to_excel(yeni_dosya_adi, index=False)
        

        # İndirilen dosyaları silme
        for i in range(1, 11):
            file_name = f"link_{i}.xlsx"
            if os.path.exists(file_name):
                os.remove(file_name)
               

    if __name__ == "__main__":
        asyncio.run(main())




















user_input = "E"


google_sheet_url = "https://docs.google.com/spreadsheets/d/1FJwRFD6ikSsy3uGFRiKp94Iaj1Jd5xerEzJfxJgS1f8/gviz/tq?tqx=out:csv"

try:
    google_df = pd.read_csv(google_sheet_url)
    google_excel_file = "Hariç Tutulacak Sipariş Numaraları.xlsx"
    google_df.to_excel(google_excel_file, index=False)
except requests.exceptions.RequestException as e:
    pass


def main():
    while True:      
        if user_input == "E":
            # İki Excel dosyasını okuyun
            birlesik_excel = pd.read_excel("birlesik_excel.xlsx")
            haric_excel = pd.read_excel("Hariç Tutulacak Sipariş Numaraları.xlsx")

            # İlk sütunlara göre birleşik Excel verisinden, haric Excel verisinde bulunan satırları filtreleyin
            birlesik_excel = birlesik_excel[~birlesik_excel['Id'].isin(haric_excel['Id'])]

            # Sonucu mevcut "birlesik_excel.xlsx" dosyasına kaydedin (var olan dosyanın üstüne yazacak)
            birlesik_excel.to_excel("birlesik_excel.xlsx", index=False)
            break
        elif user_input == "H":
            break
        else:
            print("Geçerli bir seçenek giriniz (E/H).")

if __name__ == "__main__":
    main()




file_path = "Hariç Tutulacak Sipariş Numaraları.xlsx"



if os.path.exists(file_path):
    os.remove(file_path)













    







    









    google_sheet_url = "https://docs.google.com/spreadsheets/d/1PgldjEkmmjLPrG9dqvaou481m9QajCOlGxa7wCjwTAQ/gviz/tq?tqx=out:csv"

    try:
        google_df = pd.read_csv(google_sheet_url)
        google_excel_file = "Kara Liste.xlsx"
        google_df.to_excel(google_excel_file, index=False)
    except requests.exceptions.RequestException as e:
        pass
       







    # "birlesik_excel" dosyasını yükle
    birlesik_excel_file = "birlesik_excel.xlsx"
    birlesik_df = pd.read_excel(birlesik_excel_file)

    # "Kara Liste" dosyasını yükle
    kara_liste_file = "Kara Liste.xlsx"
    kara_liste_df = pd.read_excel(kara_liste_file)

    # Çıkarılan satırları tutmak için bir liste oluştur
    cikarilan_satirlar = []

    # Kara listedeki telefon numaralarını içeren satırları bul, çıkar ve ayrı bir liste'e ekle
    for telefon in kara_liste_df.iloc[:, 0]:
        matching_rows = birlesik_df[birlesik_df["TeslimatTelefon"] == telefon]
        
        # ÖdemeTipi sütunu "Kapıda Ödeme" ise işlem yap
        for index, row in matching_rows.iterrows():
            if row["OdemeTipi"] == "Kapıda Ödeme" and row["SiparisDurumu"] != "Arandı / Onaylandı":
                cikarilan_satirlar.append(row)
                birlesik_df.drop(index, inplace=True)

    # Çıkarılan satırları içeren bir DataFrame oluştur
    cikarilan_satirlar_df = pd.DataFrame(cikarilan_satirlar, columns=birlesik_df.columns)

    # Çıkarılan satırları "cikarilan_satirlar.xlsx" dosyasına kaydet
    cikarilan_satirlar_df.to_excel("cikarilan_satirlar.xlsx", index=False)

    # Güncellenmiş "birlesik_excel" verilerini aynı dosyaya kaydet (üzerine yaz)
    birlesik_df.to_excel(birlesik_excel_file, index=False)









    # "cikarilan_satirlar" dosyasını yükle
    cikarilan_satirlar_file = "cikarilan_satirlar.xlsx"
    cikarilan_satirlar_df = pd.read_excel(cikarilan_satirlar_file)

    # "Id" ve "TeslimatTelefon" sütunları hariç diğer tüm sütunları sil
    sadece_id_telefon_df = cikarilan_satirlar_df[["Id", "TeslimatTelefon"]]
    cikarilan_satirlar_df = sadece_id_telefon_df.copy()

    # Yenilenen değerleri kaldır (sadece benzersiz satırları bırak)
    cikarilan_satirlar_df.drop_duplicates(inplace=True)

    # "TeslimatTelefon" sütununu sayı biçimine çevir (formatlama)
    cikarilan_satirlar_df["TeslimatTelefon"] = cikarilan_satirlar_df["TeslimatTelefon"].apply(lambda x: '{:.0f}'.format(x))

    # Güncellenmiş verileri tekrar kaydet (sadece "Id" ve "TeslimatTelefon" sütunları olacak şekilde)
    cikarilan_satirlar_df.to_excel(cikarilan_satirlar_file, index=False)

    # "Kara Liste" dosyasını sil
    kara_liste_file = "Kara Liste.xlsx"
    if os.path.exists(kara_liste_file):
        os.remove(kara_liste_file)
       

    # "cikarilan_satirlar" dosyasının adını "Kara Liste Siparişleri.xlsx" olarak değiştir
    yeni_ad = "Kara Liste Siparişleri.xlsx"
    os.rename(cikarilan_satirlar_file, yeni_ad)
    








    google_sheet_url = "https://docs.google.com/spreadsheets/d/1iKtTb2X28bYDarV3y8YUiRWCnov8zgJBxRVuQOygrNQ/gviz/tq?tqx=out:csv"

    try:
        google_df = pd.read_csv(google_sheet_url)
        google_excel_file = "Hariç Tutulacak Sipariş Numaraları.xlsx"
        google_df.to_excel(google_excel_file, index=False)
    except requests.exceptions.RequestException as e:
        pass



    def main():
        while True:      
            
            # İki Excel dosyasını okuyun
            birlesik_excel = pd.read_excel("birlesik_excel.xlsx")
            haric_excel = pd.read_excel("Hariç Tutulacak Sipariş Numaraları.xlsx")

            # İlk sütunlara göre birleşik Excel verisinden, haric Excel verisinde bulunan satırları filtreleyin
            birlesik_excel = birlesik_excel[~birlesik_excel['Id'].isin(haric_excel['Id'])]

            # Sonucu mevcut "birlesik_excel.xlsx" dosyasına kaydedin (var olan dosyanın üstüne yazacak)
            birlesik_excel.to_excel("birlesik_excel.xlsx", index=False)
            break
    if __name__ == "__main__":
        main()


    # Kaynak dosya adı
    kaynak_excel = "birlesik_excel.xlsx"

    # Kopya dosya adı (istediğiniz adı ve konumu belirtin)
    kopya_excel = "Hazırlanan Sipariş Numaraları3.xlsx"

    # Dosyayı kopyala
    shutil.copy(kaynak_excel, kopya_excel)




    file_path = "Hariç Tutulacak Sipariş Numaraları.xlsx"

    if os.path.exists(file_path):
        os.remove(file_path)































    # Excel dosyasını okuma
    df = pd.read_excel("birlesik_excel.xlsx")


    # İşlemi gerçekleştiren fonksiyon
    def duplicate_rows(row):
        count = int(row["Adet"])
        return pd.concat([row] * count, axis=1).T

    # Tüm satırları işleme tabi tutma
    new_rows = df.apply(duplicate_rows, axis=1)

    # Yeni veri çerçevesini oluşturma
    new_df = pd.concat(new_rows.tolist(), ignore_index=True)

    # Sadece belirtilen sütunları seçme
    selected_columns = ["Id", "Barkod", "UrunAdi", "Varyant"]
    new_df = new_df[selected_columns]

    # Veriyi yeni bir Excel dosyasına yazma
    new_df.to_excel("birlesik_excel.xlsx", index=False)













    url = "https://haydigiy.online/Products/rafkodlari.php"
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")
    table = soup.find("table")
    data = []
    for row in table.find_all("tr"):
        row_data = []
        for cell in row.find_all(["th", "td"]):
            row_data.append(cell.get_text(strip=True))
        data.append(row_data)
    df = pd.DataFrame(data[1:], columns=data[0])
    df.to_excel("Raf Kodu.xlsx", index=False)
       






    # "birlesik_excel.xlsx" ve "Raf Kodu.xlsx" dosyalarını okuma
    sonuc_df = pd.read_excel("birlesik_excel.xlsx")
    google_sheet_df = pd.read_excel("Raf Kodu.xlsx")

    # "birlesik_excel.xlsx" dosyasına yeni bir sütun ekleyerek başlangıçta "Raf Kodu Yok" değerleri ile doldurma
    sonuc_df["GoogleSheetVerisi"] = "Raf Kodu Yok"

    # Her bir "Barkod" değeri için işlem yapma
    for index, row in sonuc_df.iterrows():
        barkod = row["Barkod"]
        
        # "Raf Kodu.xlsx" dosyasında ilgili "Barkod"u arama
        matching_row = google_sheet_df[google_sheet_df.iloc[:, 0] == barkod]
        
        # Eşleşen "Barkod" varsa ve karşılık gelen hücre boş değilse, değeri "GoogleSheetVerisi" sütununa yazma
        if not matching_row.empty and not pd.isnull(matching_row.iloc[0, 2]):
            sonuc_df.at[index, "GoogleSheetVerisi"] = matching_row.iloc[0, 2]

    # "birlesik_excel.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("birlesik_excel.xlsx", index=False)
   






    # "birlesik_excel.xlsx" dosyasını güncelleme
    sonuc_df["GoogleSheetVerisi Kopya"] = sonuc_df["GoogleSheetVerisi"]  # "GoogleSheetVerisi" sütununu kopyala
    sonuc_df["GoogleSheetVerisi Kopya"] = sonuc_df["GoogleSheetVerisi Kopya"].str.split("-", n=1).str[0]  # "-" den sonrasını temizle
    sonuc_df["GoogleSheetVerisi Kopya"] = pd.to_numeric(sonuc_df["GoogleSheetVerisi Kopya"], errors="coerce")  # Sayıya dönüştür
    sonuc_df = sonuc_df.sort_values(by="GoogleSheetVerisi Kopya")  # "GoogleSheetVerisi Kopya" sütununa göre sırala


    sonuc_df.to_excel("birlesik_excel.xlsx", index=False)














    # "birlesik_excel.xlsx" ve "Raf Kodu.xlsx" dosyalarını okuma
    sonuc_df = pd.read_excel("birlesik_excel.xlsx")
    google_sheet_df = pd.read_excel("Raf Kodu.xlsx")

    # "birlesik_excel.xlsx" dosyasına yeni bir sütun ekleyerek başlangıçta "Raf Kodu Yok" değerleri ile doldurma
    sonuc_df["Kategori"] = "Raf Kodu Yok"

    # Her bir "Barkod" değeri için işlem yapma
    for index, row in sonuc_df.iterrows():
        barkod = row["Barkod"]
        
        # "Raf Kodu.xlsx" dosyasında ilgili "Barkod"u arama
        matching_row = google_sheet_df[google_sheet_df.iloc[:, 0] == barkod]
        
        # Eşleşen "Barkod" varsa ve karşılık gelen hücre boş değilse, değeri "GoogleSheetVerisi" sütununa yazma
        if not matching_row.empty and not pd.isnull(matching_row.iloc[0, 3]):
            sonuc_df.at[index, "Kategori"] = matching_row.iloc[0, 3]

    # "birlesik_excel.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("birlesik_excel.xlsx", index=False)
 





    # Excel dosyasını oku
    excel_file_path = "birlesik_excel.xlsx"  # Excel dosyasının adını ve yolunu belirtin
    df = pd.read_excel(excel_file_path)

    # Yeni bir sütun ekleyerek işlem sonuçlarını tut
    df["Yeni Kategori"] = ""

    # İç Giyim içeren satırları işle
    innerwear_rows = df[df["Kategori"].str.contains("İç Giyim")]

    # İç Giyim içeren satırları işle
    for index, row in innerwear_rows.iterrows():
        df.loc[index, "Yeni Kategori"] = "İç Giyim"

    # Sonucu kaydet
    output_file_path = "birlesik_excel.xlsx"  # Sonucun kaydedileceği dosya adı ve yolunu belirtin
    df.to_excel(output_file_path, index=False)

 

    # Excel dosyasını oku
    excel_file_path = "birlesik_excel.xlsx"  # Excel dosyasının adını ve yolunu belirtin
    df = pd.read_excel(excel_file_path)

    # "Id" değerlerine göre grupla ve işlemi yap
    grouped = df.groupby("Id")

    # "Id" değerlerine göre "Yeni Kategori" değerini güncelle
    for group_name, group_data in grouped:
        if any(row["Yeni Kategori"] == "İç Giyim" for _, row in group_data.iterrows()):
            df.loc[df["Id"] == group_name, "Yeni Kategori"] = "İç Giyim"

    # Sonucu kaydet
    output_file_path = "birlesik_excel.xlsx"  # Sonucun kaydedileceği dosya adı ve yolunu belirtin
    df.to_excel(output_file_path, index=False)

   







    # Excel dosyasını oku
    excel_file_path = "birlesik_excel.xlsx"  # Ana Excel dosyasının adını ve yolunu belirtin
    df = pd.read_excel(excel_file_path)

    # "Yeni Kategori" değeri "İç Giyim" olan satırları seç
    innerwear_rows = df[df["Yeni Kategori"] == "İç Giyim"]

    # Ayrı Excel dosyasına kaydet
    output_file_path = "İç Giyim.xlsx"  # Ayrı dosyanın adını ve yolunu belirtin
    innerwear_rows.to_excel(output_file_path, index=False)

    # Ana DataFrame'den "İç Giyim" satırları sil
    df = df[df["Yeni Kategori"] != "İç Giyim"]
    df.drop(columns=["Yeni Kategori"], inplace=True)  # "Yeni Kategori" sütununu sil

    # Ana Excel dosyasını güncelle
    df.to_excel(excel_file_path, index=False)

    












    excel_to_delete = "Raf Kodu.xlsx"

    # Excel dosyasını sil
    if os.path.exists(excel_to_delete):
        os.remove(excel_to_delete)
       
    else:
        print("Dosya bulunamadı:", excel_to_delete)



















    old_file_path = "birlesik_excel.xlsx"
    new_file_path = "Kalanlar.xlsx"

    # Dosyanın adını değiştir
    os.rename(old_file_path, new_file_path)

    







    # Excel dosyasını oku
    excel_file_path = "Kalanlar.xlsx"  # Kalanlar Excel dosyasının adını ve yolunu belirtin
    df = pd.read_excel(excel_file_path)

    # "GoogleSheetVerisi Kopya" ve "Kategori" sütunlarını sil
    columns_to_drop = ["GoogleSheetVerisi Kopya", "Kategori"]
    df.drop(columns=columns_to_drop, inplace=True)

    # Dosyayı güncelle
    df.to_excel(excel_file_path, index=False)






    # Excel dosyasını oku
    excel_file_path = "İç Giyim.xlsx"  # Kalanlar Excel dosyasının adını ve yolunu belirtin
    df = pd.read_excel(excel_file_path)

    # "GoogleSheetVerisi Kopya" ve "Kategori" sütunlarını sil
    columns_to_drop = ["GoogleSheetVerisi Kopya", "Kategori", "Yeni Kategori"]
    df.drop(columns=columns_to_drop, inplace=True)

    # Dosyayı güncelle
    df.to_excel(excel_file_path, index=False)

    



























































    sonuc_df = pd.read_excel("Kalanlar.xlsx")

    # "Kalanlar.xlsx" dosyasını güncelleme
    sonuc_df["UrunAdi Kopya"] = sonuc_df["UrunAdi"]  # "UrunAdi" sütununu kopyala
    sonuc_df["UrunAdi Kopya"] = sonuc_df["UrunAdi Kopya"].str.split("-", n=1).str[1]  # "-" den öncesini temizle

    # "UrunAdi Kopya" sütununu "Kalanlar.xlsx" dosyasına ekleyerek güncelleme
    with pd.ExcelWriter("Kalanlar.xlsx") as writer:
        sonuc_df.to_excel(writer, index=False)

   








    # "UrunAdi" sütununu en sağına yapıştırma
    sonuc_df["UrunAdiKopya2"] = sonuc_df["UrunAdi"]

    # "UrunAdiKopya2" sütununda " - " dan sonrasını ve son boşluktan öncesini silme
    sonuc_df["UrunAdiKopya2"] = sonuc_df["UrunAdiKopya2"].apply(lambda x: x.split(" - ")[0].rsplit(" ", 1)[-1].strip() if " - " in x else x)

    # "UrunAdiKopya2" sütununu en sağa taşıma
    column_order = list(sonuc_df.columns)
    column_order.remove("UrunAdiKopya2")
    column_order.append("UrunAdiKopya2")
    sonuc_df = sonuc_df[column_order]

    # "Kalanlar.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("Kalanlar.xlsx", index=False)
    






    # "UrunAdi" sütununu tablonun sonuna bir kez daha kopyalama ve düzenleme
    sonuc_df["UrunAdiKopya3"] = sonuc_df["UrunAdi"].apply(lambda x: x.split(" - ", 1)[0].strip() if " - " in x else x)

    # "UrunAdiKopya3" sütununu en sağa taşıma
    column_order = list(sonuc_df.columns)
    column_order.remove("UrunAdiKopya3")
    column_order.append("UrunAdiKopya3")
    sonuc_df = sonuc_df[column_order]

    # "Kalanlar.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("Kalanlar.xlsx", index=False)
    









    # Verileri birleştirip yeni sütun oluşturma
    sonuc_df["BirlesikVeri"] = sonuc_df["UrunAdi Kopya"] + " - " + sonuc_df["UrunAdiKopya2"] + " - " + sonuc_df["Varyant"]

    # "BirlesikVeri" sütununu en sağa taşıma
    column_order = list(sonuc_df.columns)
    column_order.remove("BirlesikVeri")
    column_order.append("BirlesikVeri")
    sonuc_df = sonuc_df[column_order]

    # "Kalanlar.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("Kalanlar.xlsx", index=False)
  

    # "BirlesikVeri" sütunundaki "Beden:" ibaresini çıkarma
    sonuc_df["BirlesikVeri"] = sonuc_df["BirlesikVeri"].str.replace("Beden:", "")

    # "Kalanlar.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("Kalanlar.xlsx", index=False)
    





    # Belirtilen sütunları silme
    columns_to_drop = ["UrunAdi", "Varyant", "UrunAdi Kopya", "UrunAdiKopya2"]
    sonuc_df.drop(columns_to_drop, axis=1, inplace=True)

    # "Kalanlar.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("Kalanlar.xlsx", index=False)
    






    # "Id" sütununu teke düşürme
    unique_ids = sonuc_df["Id"].drop_duplicates()

    # Yeni bir Excel sayfası oluşturma ve "Id" değerlerini yazma
    with pd.ExcelWriter("Kalanlar.xlsx", engine="openpyxl", mode="a") as writer:
        unique_ids.to_excel(writer, sheet_name="Unique Ids", index=False)

    










    

    # Sonuç dosyasını yükle
    file_path = "Kalanlar.xlsx"
    wb = load_workbook(file_path)
    sheet = wb["Unique Ids"]

    # Başlangıç sütunu ve satırı
    start_column = 2
    start_row = 2

    # Toplam tekrar sayısı ve her tekrardaki numara adedi
    repeat_count = 50
    numbers_per_repeat = 28

    # Verileri ekleme
    for _ in range(repeat_count):
        for num in range(1, numbers_per_repeat + 1):
            sheet.cell(row=start_row, column=start_column, value=num)
            start_row += 1

    # Değişiklikleri kaydetme
    wb.save("Kalanlar.xlsx")

    





    # Sonuç dosyasını yükle
    file_path = "Kalanlar.xlsx"
    wb = load_workbook(file_path)
    sheet = wb["Unique Ids"]

    # Başlangıç sütunu ve satırı
    start_column = 3
    start_row = 2

    # Toplam tekrar sayısı ve her tekrardaki numara adedi
    repeat_count = 50
    numbers_per_repeat = 28

    # Verileri ekleme
    for repeat in range(repeat_count):
        for num in range(1, numbers_per_repeat + 1):
            sheet.cell(row=start_row, column=start_column, value=(repeat % numbers_per_repeat) + 1)
            start_row += 1

    # Değişiklikleri kaydetme
    wb.save("Kalanlar.xlsx")










    # Sonuç dosyasını yükle
    file_path = "Kalanlar.xlsx"
    wb = load_workbook(file_path)
    unique_ids_sheet = wb["Unique Ids"]
    main_sheet = wb["Sheet1"]

    # "Id" sütununun verilerini al
    id_column = main_sheet["A"][1:]
    unique_ids_column = unique_ids_sheet["A"][1:]

    # Karşılık gelen değerleri bulup "Sheet1" sayfasının en sağında yeni bir sütuna ekle
    new_column = main_sheet.max_column + 1
    main_sheet.cell(row=1, column=new_column, value="Matching Value")

    for id_cell in id_column:
        id_value = id_cell.value
        for unique_id_cell in unique_ids_column:
            if unique_id_cell.value == id_value:
                matching_value = unique_id_cell.offset(column=1).value
                main_sheet.cell(row=id_cell.row, column=new_column, value=matching_value)
                break

    # Değişiklikleri kaydetme
    wb.save("Kalanlar.xlsx")

    










    # Sonuç dosyasını yükle
    file_path = "Kalanlar.xlsx"
    wb = load_workbook(file_path)
    unique_ids_sheet = wb["Unique Ids"]
    main_sheet = wb["Sheet1"]

    # "Id" sütununun verilerini al
    id_column = main_sheet["A"][1:]
    unique_ids_column = unique_ids_sheet["A"][1:]

    # Karşılık gelen değerleri bulup "Sheet1" sayfasının en sağında yeni bir sütuna ekle
    new_column = main_sheet.max_column + 1
    main_sheet.cell(row=1, column=new_column, value="Matching Value (3rd Column)")

    for id_cell in id_column:
        id_value = id_cell.value
        for unique_id_cell in unique_ids_column:
            if unique_id_cell.value == id_value:
                matching_value = unique_id_cell.offset(column=2).value
                main_sheet.cell(row=id_cell.row, column=new_column, value=matching_value)
                break

    # Değişiklikleri kaydetme
    wb.save("Kalanlar.xlsx")

    









    # Sonuç dosyasını yükle
    file_path = "Kalanlar.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Sütun başlıklarını değiştir
    new_column_titles = {
        "Id": "SiparişNO",
        "BirlesikVeri": "ÜRÜN",
        "GoogleSheetVerisi": "RAF KODU",
        "UrunAdiKopya3": "ÜRÜN ADI",
        "Matching Value": "KUTU",
        "Matching Value (3rd Column)": "ÇN"
    }

    for col_idx, col_name in enumerate(main_sheet.iter_cols(min_col=1, max_col=main_sheet.max_column), start=1):
        old_title = col_name[0].value
        new_title = new_column_titles.get(old_title, old_title)
        col_name[0].value = new_title

    # Değişiklikleri kaydetme
    wb.save("Kalanlar.xlsx")

    








    

    # Sonuç dosyasını yükle
    file_path = "Kalanlar.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Sütunların yeni sıralaması
    new_column_order = [
        "RAF KODU",
        "ÜRÜN",
        "Barkod",
        "KUTU",
        "ÜRÜN ADI",
        "ÇN",
        "SiparişNO"
    ]

    # Yeni bir DataFrame oluştur
    data = main_sheet.iter_rows(min_row=2, values_only=True)
    df = pd.DataFrame(data, columns=[cell.value for cell in main_sheet[1]])

    # Sütunları yeni sıralamaya göre düzenle
    df = df[new_column_order]

    # Mevcut başlıkları güncelle
    for idx, column_name in enumerate(new_column_order, start=1):
        main_sheet.cell(row=1, column=idx, value=column_name)

    # DataFrame verilerini sayfaya yaz
    for r_idx, row in enumerate(df.values, 2):
        for c_idx, value in enumerate(row, 1):
            main_sheet.cell(row=r_idx, column=c_idx, value=value)

    # Değişiklikleri kaydet
    wb.save("Kalanlar.xlsx")

   







    

    # Sonuç dosyasını yükle
    file_path = "Kalanlar.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Hücreleri ortala ve ortaya hizala
    for row in main_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")


    # Değişiklikleri kaydet
    wb.save("Kalanlar.xlsx")

  






    
    

    # Sonuç dosyasını yükle
    file_path = "Kalanlar.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Kenarlık stili oluştur
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Hücreleri kalın yap, yazı fontunu 14 yap ve kenarlık ekle
    for row in main_sheet.iter_rows():
        for cell in row:
            cell.font = Font(bold=True, size=14)
            cell.border = border_style

    # Değişiklikleri kaydet
    wb.save("Kalanlar.xlsx")



    

    # Sonuç dosyasını yükle
    file_path = "Kalanlar.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # "RAF KODU" sütununu 45 piksel yap
    main_sheet.column_dimensions["C"].width = 45

    # Tüm hücreleri en uygun sütun genişliği olarak ayarla
    for column in main_sheet.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        main_sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Değişiklikleri kaydet
    wb.save("Kalanlar.xlsx")




    # Sonuç dosyasını yükle
    file_path = "Kalanlar.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # İlk sütunu (A sütunu) 45 piksel genişliğinde yap
    main_sheet.column_dimensions["A"].width = 45
    main_sheet.column_dimensions["C"].width = 14
    main_sheet.column_dimensions["G"].width = 14
    main_sheet.column_dimensions["D"].width = 9
    main_sheet.column_dimensions["F"].width = 5

    # Değişiklikleri kaydet
    wb.save("Kalanlar.xlsx")









    

    # Sonuç dosyasını yükle
    file_path = "Kalanlar.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Tüm hücrelere "Metni Kaydır" formatını uygula
    for row in main_sheet.iter_rows():
        for cell in row:
            new_alignment = copy(cell.alignment)
            new_alignment.wrap_text = True
            cell.alignment = new_alignment

    # Değişiklikleri kaydet
    wb.save("Kalanlar.xlsx")








    

    # Sonuç dosyasını yükle
    file_path = "Kalanlar.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Tabloyu oluşturma
    table = Table(displayName="MyTable", ref=main_sheet.dimensions)

    # Tablo stili oluşturma (gri-beyaz)
    style = TableStyleInfo(
        name="TableStyleLight1", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True
    )

    # Tabloya stil atama
    table.tableStyleInfo = style

    # Tabloyu sayfaya ekleme
    main_sheet.add_table(table)

    # Değişiklikleri kaydetme
    wb.save("Kalanlar.xlsx")



    

    def create_bat_files(data, output_folder, batch_size=28):
        batch_count = 1
        batch_data = []
        remaining_data = data

        while len(remaining_data) > 0:
            current_batch = remaining_data[:batch_size]
            batch_data.extend(current_batch)

            bat_file_path = os.path.join(output_folder, f"BAT{batch_count}.bat")
            with open(bat_file_path, "w") as file:
                link = f'start "" https://task.haydigiy.com/admin/order/printorder/?orderId={current_batch[0]}&isPdf=False\n'
                file.write(link)
                file.write('timeout -t 1\n')  # Add the timeout line

                for value in current_batch[1:]:
                    link = f"https://task.haydigiy.com/admin/order/printorder/?orderId={value}&isPdf=False"
                    file.write(f'start "" {link}\n')

            batch_data = []
            remaining_data = remaining_data[batch_size:]
            batch_count += 1

    # Klasör oluştur
    output_folder = "Kalanlar"
    os.makedirs(output_folder, exist_ok=True)

    # Sonuç dosyasını yükle
    file_path = "Kalanlar.xlsx"
    wb = load_workbook(file_path)
    unique_ids_sheet = wb["Unique Ids"]

    # "Id" sütunundaki verileri al
    id_column = unique_ids_sheet["A"][1:]

    # Verileri bir listeye dönüştür
    id_values = [cell.value for cell in id_column if cell.value is not None]

    # .bat dosyalarını oluştur ve klasöre kaydet
    create_bat_files(id_values, output_folder)

    # Excel dosyasını klasöre taşı
    shutil.copy(file_path, os.path.join(output_folder, "Kalanlar.xlsx"))

    gc.collect()

    # Klasör dışında kalan Excel dosyasını sil
    os.remove(file_path)





















































    sonuc_df = pd.read_excel("İç Giyim.xlsx")

    # "İç Giyim.xlsx" dosyasını güncelleme
    sonuc_df["UrunAdi Kopya"] = sonuc_df["UrunAdi"]  # "UrunAdi" sütununu kopyala
    sonuc_df["UrunAdi Kopya"] = sonuc_df["UrunAdi Kopya"].str.split("-", n=1).str[1]  # "-" den öncesini temizle

    # "UrunAdi Kopya" sütununu "İç Giyim.xlsx" dosyasına ekleyerek güncelleme
    with pd.ExcelWriter("İç Giyim.xlsx") as writer:
        sonuc_df.to_excel(writer, index=False)









    # "UrunAdi" sütununu en sağına yapıştırma
    sonuc_df["UrunAdiKopya2"] = sonuc_df["UrunAdi"]

    # "UrunAdiKopya2" sütununda " - " dan sonrasını ve son boşluktan öncesini silme
    sonuc_df["UrunAdiKopya2"] = sonuc_df["UrunAdiKopya2"].apply(lambda x: x.split(" - ")[0].rsplit(" ", 1)[-1].strip() if " - " in x else x)

    # "UrunAdiKopya2" sütununu en sağa taşıma
    column_order = list(sonuc_df.columns)
    column_order.remove("UrunAdiKopya2")
    column_order.append("UrunAdiKopya2")
    sonuc_df = sonuc_df[column_order]

    # "İç Giyim.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("İç Giyim.xlsx", index=False)






    # "UrunAdi" sütununu tablonun sonuna bir kez daha kopyalama ve düzenleme
    sonuc_df["UrunAdiKopya3"] = sonuc_df["UrunAdi"].apply(lambda x: x.split(" - ", 1)[0].strip() if " - " in x else x)

    # "UrunAdiKopya3" sütununu en sağa taşıma
    column_order = list(sonuc_df.columns)
    column_order.remove("UrunAdiKopya3")
    column_order.append("UrunAdiKopya3")
    sonuc_df = sonuc_df[column_order]

    # "İç Giyim.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("İç Giyim.xlsx", index=False)









    # Verileri birleştirip yeni sütun oluşturma
    sonuc_df["BirlesikVeri"] = sonuc_df["UrunAdi Kopya"] + " - " + sonuc_df["UrunAdiKopya2"] + " - " + sonuc_df["Varyant"]

    # "BirlesikVeri" sütununu en sağa taşıma
    column_order = list(sonuc_df.columns)
    column_order.remove("BirlesikVeri")
    column_order.append("BirlesikVeri")
    sonuc_df = sonuc_df[column_order]

    # "İç Giyim.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("İç Giyim.xlsx", index=False)

    # "BirlesikVeri" sütunundaki "Beden:" ibaresini çıkarma
    sonuc_df["BirlesikVeri"] = sonuc_df["BirlesikVeri"].str.replace("Beden:", "")

    # "İç Giyim.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("İç Giyim.xlsx", index=False)





    # Belirtilen sütunları silme
    columns_to_drop = ["UrunAdi", "Varyant", "UrunAdi Kopya", "UrunAdiKopya2"]
    sonuc_df.drop(columns_to_drop, axis=1, inplace=True)

    # "İç Giyim.xlsx" dosyasını güncelleme
    sonuc_df.to_excel("İç Giyim.xlsx", index=False)






    # "Id" sütununu teke düşürme
    unique_ids = sonuc_df["Id"].drop_duplicates()

    # Yeni bir Excel sayfası oluşturma ve "Id" değerlerini yazma
    with pd.ExcelWriter("İç Giyim.xlsx", engine="openpyxl", mode="a") as writer:
        unique_ids.to_excel(writer, sheet_name="Unique Ids", index=False)











    

    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    sheet = wb["Unique Ids"]

    # Başlangıç sütunu ve satırı
    start_column = 2
    start_row = 2

    # Toplam tekrar sayısı ve her tekrardaki numara adedi
    repeat_count = 50
    numbers_per_repeat = 28

    # Verileri ekleme
    for _ in range(repeat_count):
        for num in range(1, numbers_per_repeat + 1):
            sheet.cell(row=start_row, column=start_column, value=num)
            start_row += 1

    # Değişiklikleri kaydetme
    wb.save("İç Giyim.xlsx")






    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    sheet = wb["Unique Ids"]

    # Başlangıç sütunu ve satırı
    start_column = 3
    start_row = 2

    # Toplam tekrar sayısı ve her tekrardaki numara adedi
    repeat_count = 50
    numbers_per_repeat = 28

    # Verileri ekleme
    for repeat in range(repeat_count):
        for num in range(1, numbers_per_repeat + 1):
            sheet.cell(row=start_row, column=start_column, value=(repeat % numbers_per_repeat) + 1)
            start_row += 1

    # Değişiklikleri kaydetme
    wb.save("İç Giyim.xlsx")










    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    unique_ids_sheet = wb["Unique Ids"]
    main_sheet = wb["Sheet1"]

    # "Id" sütununun verilerini al
    id_column = main_sheet["A"][1:]
    unique_ids_column = unique_ids_sheet["A"][1:]

    # Karşılık gelen değerleri bulup "Sheet1" sayfasının en sağında yeni bir sütuna ekle
    new_column = main_sheet.max_column + 1
    main_sheet.cell(row=1, column=new_column, value="Matching Value")

    for id_cell in id_column:
        id_value = id_cell.value
        for unique_id_cell in unique_ids_column:
            if unique_id_cell.value == id_value:
                matching_value = unique_id_cell.offset(column=1).value
                main_sheet.cell(row=id_cell.row, column=new_column, value=matching_value)
                break

    # Değişiklikleri kaydetme
    wb.save("İç Giyim.xlsx")











    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    unique_ids_sheet = wb["Unique Ids"]
    main_sheet = wb["Sheet1"]

    # "Id" sütununun verilerini al
    id_column = main_sheet["A"][1:]
    unique_ids_column = unique_ids_sheet["A"][1:]

    # Karşılık gelen değerleri bulup "Sheet1" sayfasının en sağında yeni bir sütuna ekle
    new_column = main_sheet.max_column + 1
    main_sheet.cell(row=1, column=new_column, value="Matching Value (3rd Column)")

    for id_cell in id_column:
        id_value = id_cell.value
        for unique_id_cell in unique_ids_column:
            if unique_id_cell.value == id_value:
                matching_value = unique_id_cell.offset(column=2).value
                main_sheet.cell(row=id_cell.row, column=new_column, value=matching_value)
                break

    # Değişiklikleri kaydetme
    wb.save("İç Giyim.xlsx")










    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Sütun başlıklarını değiştir
    new_column_titles = {
        "Id": "SiparişNO",
        "BirlesikVeri": "ÜRÜN",
        "GoogleSheetVerisi": "RAF KODU",
        "UrunAdiKopya3": "ÜRÜN ADI",
        "Matching Value": "KUTU",
        "Matching Value (3rd Column)": "ÇN"
    }

    for col_idx, col_name in enumerate(main_sheet.iter_cols(min_col=1, max_col=main_sheet.max_column), start=1):
        old_title = col_name[0].value
        new_title = new_column_titles.get(old_title, old_title)
        col_name[0].value = new_title

    # Değişiklikleri kaydetme
    wb.save("İç Giyim.xlsx")









    

    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Sütunların yeni sıralaması
    new_column_order = [
        "RAF KODU",
        "ÜRÜN",
        "Barkod",
        "KUTU",
        "ÜRÜN ADI",
        "ÇN",
        "SiparişNO"
    ]

    # Yeni bir DataFrame oluştur
    data = main_sheet.iter_rows(min_row=2, values_only=True)
    df = pd.DataFrame(data, columns=[cell.value for cell in main_sheet[1]])

    # Sütunları yeni sıralamaya göre düzenle
    df = df[new_column_order]

    # Mevcut başlıkları güncelle
    for idx, column_name in enumerate(new_column_order, start=1):
        main_sheet.cell(row=1, column=idx, value=column_name)

    # DataFrame verilerini sayfaya yaz
    for r_idx, row in enumerate(df.values, 2):
        for c_idx, value in enumerate(row, 1):
            main_sheet.cell(row=r_idx, column=c_idx, value=value)

    # Değişiklikleri kaydet
    wb.save("İç Giyim.xlsx")








    

    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Hücreleri ortala ve ortaya hizala
    for row in main_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")


    # Değişiklikleri kaydet
    wb.save("İç Giyim.xlsx")







    
    

    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Kenarlık stili oluştur
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Hücreleri kalın yap, yazı fontunu 14 yap ve kenarlık ekle
    for row in main_sheet.iter_rows():
        for cell in row:
            cell.font = Font(bold=True, size=14)
            cell.border = border_style

    # Değişiklikleri kaydet
    wb.save("İç Giyim.xlsx")



    

    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # "RAF KODU" sütununu 45 piksel yap
    main_sheet.column_dimensions["C"].width = 45

    # Tüm hücreleri en uygun sütun genişliği olarak ayarla
    for column in main_sheet.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        main_sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # Değişiklikleri kaydet
    wb.save("İç Giyim.xlsx")




    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # İlk sütunu (A sütunu) 45 piksel genişliğinde yap
    main_sheet.column_dimensions["A"].width = 45
    main_sheet.column_dimensions["C"].width = 14
    main_sheet.column_dimensions["G"].width = 14
    main_sheet.column_dimensions["D"].width = 9
    main_sheet.column_dimensions["F"].width = 5

    # Değişiklikleri kaydet
    wb.save("İç Giyim.xlsx")









    

    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Tüm hücrelere "Metni Kaydır" formatını uygula
    for row in main_sheet.iter_rows():
        for cell in row:
            new_alignment = copy(cell.alignment)
            new_alignment.wrap_text = True
            cell.alignment = new_alignment

    # Değişiklikleri kaydet
    wb.save("İç Giyim.xlsx")








    

    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    main_sheet = wb["Sheet1"]

    # Tabloyu oluşturma
    table = Table(displayName="MyTable", ref=main_sheet.dimensions)

    # Tablo stili oluşturma (gri-beyaz)
    style = TableStyleInfo(
        name="TableStyleLight1", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True
    )

    # Tabloya stil atama
    table.tableStyleInfo = style

    # Tabloyu sayfaya ekleme
    main_sheet.add_table(table)

    # Değişiklikleri kaydetme
    wb.save("İç Giyim.xlsx")



    

    def create_bat_files(data, output_folder, batch_size=28):
        batch_count = 1
        batch_data = []
        remaining_data = data

        while len(remaining_data) > 0:
            current_batch = remaining_data[:batch_size]
            batch_data.extend(current_batch)

            bat_file_path = os.path.join(output_folder, f"BAT{batch_count}.bat")
            with open(bat_file_path, "w") as file:
                link = f'start "" https://task.haydigiy.com/admin/order/printorder/?orderId={current_batch[0]}&isPdf=False\n'
                file.write(link)
                file.write('timeout -t 1\n')  # Add the timeout line

                for value in current_batch[1:]:
                    link = f"https://task.haydigiy.com/admin/order/printorder/?orderId={value}&isPdf=False"
                    file.write(f'start "" {link}\n')

            batch_data = []
            remaining_data = remaining_data[batch_size:]
            batch_count += 1

    # Klasör oluştur
    output_folder = "İç Giyim"
    os.makedirs(output_folder, exist_ok=True)

    # Sonuç dosyasını yükle
    file_path = "İç Giyim.xlsx"
    wb = load_workbook(file_path)
    unique_ids_sheet = wb["Unique Ids"]

    # "Id" sütunundaki verileri al
    id_column = unique_ids_sheet["A"][1:]

    # Verileri bir listeye dönüştür
    id_values = [cell.value for cell in id_column if cell.value is not None]

    # .bat dosyalarını oluştur ve klasöre kaydet
    create_bat_files(id_values, output_folder)

    # Excel dosyasını klasöre taşı
    shutil.copy(file_path, os.path.join(output_folder, "İç Giyim.xlsx"))

    gc.collect()

    # Klasör dışında kalan Excel dosyasını sil
    os.remove(file_path)


































    
    
 

    # Klasör adları
    folders = ["İç Giyim", "Kalanlar"]

    # Bugünkü tarihi al
    current_date = datetime.datetime.now().strftime("%Y-%m-%d")

    # Oluşturulacak zip dosyasının adı
    zip_filename = "Faturasız Siparişler.zip"

    # Klasörleri kontrol et ve gerektiğinde sil veya zip'e ekle
    with zipfile.ZipFile(zip_filename, 'w') as zipf:
        for folder in folders:
            folder_path = os.path.join(".", folder)
            folder_contents = os.listdir(folder_path)
            bat_files = [file for file in folder_contents if file.endswith(".bat")]

            if bat_files:
                for root, _, files in os.walk(folder_path):
                    for file in files:
                        file_path = os.path.join(root, file)
                        zipf.write(file_path, os.path.relpath(file_path, "."))
                
                
            else:
                for root, dirs, files in os.walk(folder_path, topdown=False):
                    for file in files:
                        file_path = os.path.join(root, file)
                        os.remove(file_path)
                    for dir in dirs:
                        dir_path = os.path.join(root, dir)
                        os.rmdir(dir_path)
                os.rmdir(folder_path)





    # Klasörleri sil
    for folder in folders:
        folder_path = os.path.join(".", folder)
        if os.path.exists(folder_path):
            shutil.rmtree(folder_path)


    print(Fore.YELLOW + "Faturasız Siparişler Hazırlandı")

else:
    pass

#endregion






# Tanımlanan Excel dosyalarının listesi
excel_files = [
    "Hazırlanan Sipariş Numaraları.xlsx",
    "Hazırlanan Sipariş Numaraları2.xlsx",
    "Hazırlanan Sipariş Numaraları3.xlsx",
    "Hazırlanan Sipariş Numaraları4.xlsx",
    "Hazırlanan Sipariş Numaraları5.xlsx"
]

# Boş bir DataFrame oluştur
merged_data = pd.DataFrame()

# Excel dosyalarını oku ve birleştir
for excel_file in excel_files:
    if os.path.exists(excel_file):
        data = pd.read_excel(excel_file)
        merged_data = pd.concat([merged_data, data], ignore_index=True)

        # Mevcut Excel dosyasını sil
        os.remove(excel_file)

# Birleştirilmiş veriyi yeni Excel dosyasına yaz
merged_data.to_excel("Hazırlanan Sipariş Numaraları.xlsx", index=False)






# Birleştirilmiş Excel dosyasını oku
merged_data = pd.read_excel("Hazırlanan Sipariş Numaraları.xlsx")

# "Id" dışındaki tüm sütunları sil
merged_data = merged_data[['Id']]


# Güncellenmiş veriyi Excel dosyasına yaz
merged_data.to_excel("Hazırlanan Sipariş Numaraları.xlsx", index=False)








# Excel dosyasını oku
merged_data = pd.read_excel("Hazırlanan Sipariş Numaraları.xlsx")

# Benzersiz değerleri içeren DataFrame'i oluştur
unique_data = merged_data.drop_duplicates()

# Güncellenmiş veriyi Excel dosyasına yaz
unique_data.to_excel("Hazırlanan Sipariş Numaraları.xlsx", index=False)




# Boş dosyaları kontrol etme ve silme işlemi
for dosya in ["Kara Liste Siparişleri.xlsx", "Çift Siparişler.xlsx", "2500 TL Üzeri Aranacak Siparişler.xlsx", "2500 TL Üzeri Çift Siparişler.xlsx"]:
    try:
        df = pd.read_excel(dosya) 
        if df.dropna().empty:  
            os.remove(dosya)
        else:
            pass
    except FileNotFoundError:
        print(f"'{dosya}' dosyası bulunamadı.")
    except Exception as e:
        print(f"'{dosya}' dosyasını kontrol ederken bir hata oluştu: {str(e)}")


